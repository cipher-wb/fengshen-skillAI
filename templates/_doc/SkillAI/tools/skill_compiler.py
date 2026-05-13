#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SkillGraph IR → JSON 编译器 (PoC v0.1)

用法:
    python skill_compiler.py <input.skill.yaml> [--out <output.json>] [--dry-run] [-v]

输入: 符合 ir/ir_schema.json 的 YAML
输出: SkillGraph_<id>_<name>.json，可被 Unity SkillEditor 加载

PoC 阶段支持的 flow step:
  - cast_anim         播放角色动作
  - cast_effect       播放特效
  - delay             延迟执行
  - bullet (直线子弹)  创建子弹（最简版）
  - play_sound        播放音效
  - camera_shake      镜头抖动
  - apply_buff        添加 Buff
  - remove_buff       移除 Buff
  - modify_tag        修改技能参数值
  - if/then/else      条件分支
  - raw_node          直接指定 TSET_/TSCT_/TSKILLSELECT_ 节点（兜底）

依赖: pyyaml + jsonschema + skill_editor_enums.json
"""
from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import re
import sys
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
import jsonschema

TOOL_DIR = Path(__file__).parent
IR_DIR = TOOL_DIR / "ir"
IR_SCHEMA_PATH = IR_DIR / "ir_schema.json"
ENUMS_PATH = TOOL_DIR / "skill_editor_enums.json"

# 设计 Excel 路径（用于加载 BulletConfig / ModelConfig 等表格行 inline 到节点）
DESIGN_XLSX = Path(r"<<SKILL_EXCEL_PATH_WIN>>")
_TABLE_ROW_CACHE: dict[tuple[str, int], dict] = {}
_TABLE_HEADERS_CACHE: dict[str, list[str]] = {}


def load_table_row(table_name: str, row_id: int) -> dict | None:
    """从 1SkillEditor.xlsx 的指定 sheet 中读取 ID=row_id 的整行字段（英文 key→value）。

    缓存：(table_name, row_id) → dict 内存缓存避免重复读取。
    返回 None 表示未找到。
    """
    cache_key = (table_name, row_id)
    if cache_key in _TABLE_ROW_CACHE:
        return _TABLE_ROW_CACHE[cache_key]

    if not DESIGN_XLSX.exists():
        return None

    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(DESIGN_XLSX, read_only=True, data_only=True)
    if table_name not in wb.sheetnames:
        return None

    ws = wb[table_name]
    rows_iter = ws.iter_rows(values_only=True)

    # 第 1 行：中文表头
    try:
        next(rows_iter)
    except StopIteration:
        return None
    # 第 2 行：英文字段名（Key）
    try:
        en_headers = list(next(rows_iter))
    except StopIteration:
        return None
    _TABLE_HEADERS_CACHE[table_name] = [str(h) if h else "" for h in en_headers]

    # 后续：数据行，第 1 列是 ID
    for row in rows_iter:
        try:
            v0 = row[0]
            if v0 is None:
                continue
            if int(v0) == row_id:
                d: dict = {}
                for i, h in enumerate(en_headers):
                    if not h:
                        continue
                    val = row[i] if i < len(row) else None
                    d[str(h)] = _normalize_xlsx_value(val)
                _TABLE_ROW_CACHE[cache_key] = d
                return d
        except (ValueError, TypeError):
            continue

    _TABLE_ROW_CACHE[cache_key] = None  # 缓存"未找到"
    return None


def _normalize_xlsx_value(v):
    """把 xlsx 单元格值标准化为 JSON 可序列化形式。

    返回 None 表示"该字段值不可信，请跳过"——例如中文枚举字符串（"-"/"空"），
    编辑器把它反序列化为整数会失败。让 C# 类用默认值填充更安全。
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        # int 如果是浮点（xlsx 默认数字是 float），转回 int
        if isinstance(v, float) and v == int(v):
            return int(v)
        return v
    if isinstance(v, str):
        s = v.strip()
        # 布尔字符串
        if s in ("True", "true", "TRUE"):
            return True
        if s in ("False", "false", "FALSE"):
            return False
        # 真数字字符串
        if s.isdigit() or (s.startswith("-") and len(s) > 1 and s[1:].isdigit()):
            return int(s)
        # 看起来是路径/资源/英文标识符（含 / 或 _ 或纯英文）→ 保留
        if "/" in s or "_" in s:
            return s
        if s and all(c.isalnum() or c == "." for c in s) and any(c.isascii() and c.isalpha() for c in s):
            return s
        # 其他（中文枚举字符串如 "-"、"空" 等）→ 跳过
        return None
    return None


def _filter_safe_fields(row: dict) -> dict:
    """过滤掉值为 None 的字段（_normalize_xlsx_value 标记为不安全的值）"""
    return {k: v for k, v in row.items() if v is not None}


# 引用类节点的中英表名映射（→ TableManagerName）
TABLE_NAME_MAP = {
    "BulletConfig":  ("BulletConfigNode",  "TableDR.BulletConfigManager"),
    "ModelConfig":   ("ModelConfigNode",   "TableDR.ModelConfigManager"),
}

PROJECT_ROOT = Path(r"<<PROJECT_ROOT_WIN>>")
SKILL_JSON_ROOT = PROJECT_ROOT / "Assets/Thirds/NodeEditor/SkillEditor/Saves/Jsons"

# 缓存：从所有现有 SkillGraph JSON 中提取的"已知合法"引用节点 ConfigJson
# key: (cls, config_id) → ConfigJson dict
_KNOWN_REF_NODE_CACHE: dict[tuple[str, int], dict] = {}
_REF_CACHE_LOADED = False


def _ensure_known_ref_cache():
    """扫描 Saves/Jsons 下所有 SkillGraph JSON，提取 BulletConfigNode/ModelConfigNode 等引用节点，
    用作"已知合法 ConfigJson"的数据源（编辑器保存过的数据，字段完整且类型正确）。

    优先级：当编译器为 bullet_id / model_id 生成节点时，先从此缓存查找完整模板，
    找不到才降级到 xlsx 读取（含字段过滤）。
    """
    global _REF_CACHE_LOADED
    if _REF_CACHE_LOADED:
        return
    _REF_CACHE_LOADED = True

    if not SKILL_JSON_ROOT.exists():
        return

    target_classes = {"BulletConfigNode", "ModelConfigNode", "BuffConfigNode"}
    for f in SKILL_JSON_ROOT.rglob("SkillGraph_*.json"):
        if "AIGen" in f.parts:
            continue  # 跳过 AI 生成目录避免污染
        try:
            g = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        for r in g.get("references", {}).get("RefIds", []) or []:
            cls = r.get("type", {}).get("class", "")
            if cls not in target_classes:
                continue
            data = r.get("data", {})
            cj_str = data.get("ConfigJson")
            if not cj_str:
                continue
            try:
                cj = json.loads(cj_str)
            except Exception:
                continue
            cid = cj.get("ID") or data.get("ID")
            if not cid:
                continue
            cache_key = (cls, int(cid))
            if cache_key not in _KNOWN_REF_NODE_CACHE:
                _KNOWN_REF_NODE_CACHE[cache_key] = cj


def get_known_ref_config(cls: str, config_id: int) -> dict | None:
    """从缓存查"已知合法"的引用节点 ConfigJson。"""
    _ensure_known_ref_cache()
    return _KNOWN_REF_NODE_CACHE.get((cls, int(config_id)))


# 内置模板速查表（路径 + 模板根 SkillEffectID）
KNOWN_TEMPLATES = {
    "子弹通用逻辑-碰撞": {
        "path": "<<SKILLGRAPH_JSONS_ROOT>>技能模板/子弹/SkillGraph_【模板】子弹通用逻辑-碰撞.json",
        "root_effect_id": 190016404,
    },
    "子弹通用逻辑-伤害": {
        "path": "<<SKILLGRAPH_JSONS_ROOT>>技能模板/子弹/SkillGraph_【模板】子弹通用逻辑-伤害.json",
        "root_effect_id": 190016485,
    },
    "子弹通用逻辑-表现": {
        "path": "<<SKILLGRAPH_JSONS_ROOT>>技能模板/子弹/SkillGraph_【模板】子弹通用逻辑-表现.json",
        "root_effect_id": 190016523,
    },
    # v2.0 新增（M2-M6）
    "位移_按速度距离": {
        "path": "<<SKILLGRAPH_JSONS_ROOT>>技能模板/技能/SkillGraph_175_0023【模板】位移_按速度距离.json",
        "root_effect_id": 38000228,
    },
    "位移_按速度目标点": {
        "path": "<<SKILLGRAPH_JSONS_ROOT>>技能模板/技能/SkillGraph_175_0024【模板】位移_按速度目标点.json",
        "root_effect_id": 38000236,
    },
    "技能连招2段": {
        "path": "<<SKILLGRAPH_JSONS_ROOT>>技能模板/技能/SkillGraph_175_0102【模板】技能连招(2段).json",
        "root_effect_id": 38000289,
    },
    "召唤魂影单位": {
        "path": "<<SKILLGRAPH_JSONS_ROOT>>技能模板/技能/SkillGraph_146004117_【模板】召唤魂影单位.json",
        "root_effect_id": 146004117,
    },
    # v2.1 新增 — 两段式按键真机制：把当前槽位换成 second_stage SkillConfig
    "修改单位槽位": {
        "path": "<<SKILLGRAPH_JSONS_ROOT>>技能模板/技能/SkillGraph_44014633_【模板】修改单位槽位模板.json",
        "root_effect_id": 44014633,
    },
}

# 模板 TemplateParams 缓存（从目标模板 JSON 读取）
_TEMPLATE_PARAMS_CACHE: dict[str, list[dict]] = {}


def load_template_params(template_path: str) -> list[dict]:
    """从目标模板 JSON 中读出根节点的 TemplateParams 定义，用于复制到调用方节点。"""
    if template_path in _TEMPLATE_PARAMS_CACHE:
        return _TEMPLATE_PARAMS_CACHE[template_path]
    full = PROJECT_ROOT / template_path.replace("/", os.sep) if not Path(template_path).is_absolute() else Path(template_path)
    if not full.exists():
        _TEMPLATE_PARAMS_CACHE[template_path] = []
        return []
    g = json.loads(full.read_text(encoding="utf-8"))
    for r in g.get("references", {}).get("RefIds", []):
        d = r.get("data", {})
        if d.get("IsTemplate") or d.get("TemplateFlags"):
            tparams = d.get("TemplateParams", []) or []
            if tparams:
                _TEMPLATE_PARAMS_CACHE[template_path] = tparams
                return tparams
    _TEMPLATE_PARAMS_CACHE[template_path] = []
    return []

# ------------------------------------------------------------
# 加载枚举字典
# ------------------------------------------------------------
ENUMS: dict[str, Any] = {}
if ENUMS_PATH.exists():
    ENUMS = json.loads(ENUMS_PATH.read_text(encoding="utf-8")).get("enums", {})


def cn_to_int(enum_name: str, cn: str) -> int | None:
    e = ENUMS.get(enum_name)
    if not e:
        return None
    return e.get("cn_to_int", {}).get(cn)


def cn_to_enum_name(enum_name: str, cn: str) -> str | None:
    e = ENUMS.get(enum_name)
    if not e:
        return None
    return e.get("cn_to_enum", {}).get(cn)


def enum_to_int(enum_name: str, enum_str: str) -> int | None:
    """反向：用 TSET_xxx 形式查整数"""
    e = ENUMS.get(enum_name)
    if not e:
        return None
    # entries 数组
    for entry in e.get("entries", []):
        if entry["enum"] == enum_str:
            return entry["value"]
    return None


# ------------------------------------------------------------
# Reference 解析（IR 字符串 → TParam {Value, ParamType, Factor}）
# ------------------------------------------------------------
@dataclass
class TParam:
    value: int = 0
    param_type: int = 0
    factor: int = 0

    def to_dict(self) -> dict:
        # 强制转 int 防止 Python bool 透出（Python isinstance(False,int)==True 的坑）
        # C# SkillEffectConfig.Params[i].Value 是 int 字段，json bool 反序列化会失败 → Config=null → GetConfigID=0 → 编辑器报"ID异常"
        return {
            "Value":     int(self.value)      if isinstance(self.value, bool)      else self.value,
            "ParamType": int(self.param_type) if isinstance(self.param_type, bool) else self.param_type,
            "Factor":    int(self.factor)     if isinstance(self.factor, bool)     else self.factor,
        }


def resolve_ref(ref: Any) -> TParam:
    """解析 IR 中的引用字符串或纯数字 → TParam。

    IR 字符串语法：
      "entity:主体"        → TPT_COMMON_PARAM (5), TCommonParamType[主体单位实例ID]=1
      "entity:目标"        → 5, 2
      "entity:施法者"      → 5, 3
      "entity:主体伤害归属"→ 5, 4
      "attr:位置X"         → TPT_ATTR (1), TBattleNatureEnum[位置X]=59
      "attr:面向"          → 1, 91
      "tag:<id>"           → TPT_SKILL_PARAM (3), Value=<id>
      "effect_return:<id>" → TPT_FUNCTION_RETURN (2), Value=<id>
      "extra:<idx>"        → TPT_EXTRA_PARAM (4), Value=<idx>
      整数 / 数字字符串    → TPT_NULL (0)
    """
    if isinstance(ref, TParam):
        return ref
    # bool 必须先于 int 匹配（Python 中 bool 是 int 的子类）
    if isinstance(ref, bool):
        return TParam(value=1 if ref else 0)
    if isinstance(ref, int):
        return TParam(value=ref)
    if isinstance(ref, str):
        s = ref.strip()
        if s.isdigit() or (s.startswith("-") and s[1:].isdigit()):
            return TParam(value=int(s))

        if ":" in s:
            kind, val = s.split(":", 1)
            kind = kind.strip()
            val = val.strip()
            if kind == "entity":
                # 先尝试中文名，再尝试整数
                n = cn_to_int("TCommonParamType", val) if not val.isdigit() else int(val)
                # 同义别名
                aliases = {
                    "主体": "主体单位实例ID",
                    "目标": "目标单位实例ID",
                    "施法者": "施法者实例ID",
                    "主体伤害归属": "主体单位-伤害属性归属单位",
                    "目标伤害归属": "目标单位-伤害属性归属单位",
                    "施法者伤害归属": "施法者-伤害属性归属单位",
                    "施法者根": "施法者-根创建者实例ID",
                }
                if val in aliases:
                    n = cn_to_int("TCommonParamType", aliases[val])
                if n is None and val.isdigit():
                    n = int(val)
                if n is None:
                    raise CompileError(f"未知的 entity 引用: '{val}'。可选：主体/目标/施法者/主体伤害归属/施法者根")
                return TParam(value=n, param_type=5)  # TPT_COMMON_PARAM
            elif kind == "attr":
                n = cn_to_int("TBattleNatureEnum", val)
                if n is None and val.isdigit():
                    n = int(val)
                if n is None:
                    raise CompileError(f"未知的 attr 引用: '{val}'。请查 TBattleNatureEnum 枚举。")
                return TParam(value=n, param_type=1)  # TPT_ATTR
            elif kind == "tag":
                if not val.isdigit():
                    raise CompileError(f"tag 引用需要数字 ID: 'tag:{val}'")
                return TParam(value=int(val), param_type=3)  # TPT_SKILL_PARAM
            elif kind == "effect_return":
                if not val.isdigit():
                    raise CompileError(f"effect_return 引用需要数字 ID: '{val}'")
                return TParam(value=int(val), param_type=2)  # TPT_FUNCTION_RETURN
            elif kind == "extra":
                return TParam(value=int(val), param_type=4)  # TPT_EXTRA_PARAM
            else:
                raise CompileError(f"未知的引用前缀: '{kind}'。可用: entity / attr / tag / effect_return / extra")
        raise CompileError(f"无法解析的引用字符串: '{s}'")
    raise CompileError(f"无法解析的引用类型: {type(ref).__name__} 值={ref}")


# ------------------------------------------------------------
# 节点中间表示
# ------------------------------------------------------------
@dataclass
class Node:
    """编译期节点中间表示。最终序列化为 references.RefIds 中一项。"""
    rid: int = 0
    guid: str = ""
    cls: str = ""
    config_id: int = 0          # ConfigJson 中的 ID（SkillEffectConfig.ID 等）
    desc: str = ""
    config_payload: dict = field(default_factory=dict)  # 序列化进 ConfigJson
    extra_data: dict = field(default_factory=dict)      # 节点 data 顶层附加字段（如 SkillEffectType, TemplateData）
    table_name: str = ""        # 对应的表名（用于 Config2ID 与 TableTash）
    position_x: float = 0.0
    position_y: float = 0.0

    def get_config2id(self) -> str:
        return f"{self.table_name}_{self.config_id}"


# ------------------------------------------------------------
# 全局已用 ID 扫描（v2.5 / PostMortem #021）
# ------------------------------------------------------------
# 扫描所有 SkillGraph_*.json 收集 SkillEffectConfig / SkillConditionConfig /
# SkillSelectConfig / SkillTagsConfig 已用 ID，避免新分配冲突
_GLOBAL_USED_IDS_CACHE: dict[int, set[int]] = {}


def _scan_global_used_ids(exclude_skill_id: int = 0) -> set[int]:
    """扫所有 SkillGraph_*.json 收集已用 effect/cond/select/tag config ID。

    exclude_skill_id：排除指定 skill 的 .json（自身重编译场景）
    缓存：(exclude_skill_id) → set[int]
    """
    if exclude_skill_id in _GLOBAL_USED_IDS_CACHE:
        return _GLOBAL_USED_IDS_CACHE[exclude_skill_id]

    used: set[int] = set()
    pat_filename = re.compile(r"^SkillGraph_(\d+)")
    scan_dirs = [
        PROJECT_ROOT / "Assets" / "Thirds" / "NodeEditor" / "SkillEditor" / "Saves" / "Jsons",
    ]
    for d in scan_dirs:
        if not d.exists():
            continue
        for p in d.rglob("SkillGraph_*.json"):
            m = pat_filename.match(p.name)
            if m and int(m.group(1)) == exclude_skill_id:
                continue  # 排除自身
            try:
                graph = json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue
            for r in graph.get("references", {}).get("RefIds", []):
                cls = r.get("type", {}).get("class", "").split(".")[-1]
                # 仅扫"配置类节点"——TSET_ TSCT_ TSKILLSELECT_ + Tags
                relevant = (
                    cls.startswith("TSET_")
                    or cls.startswith("TSCT_")
                    or cls.startswith("TSKILLSELECT_")
                    or cls == "SkillTagsConfigNode"
                )
                if not relevant:
                    continue
                cj_str = r.get("data", {}).get("ConfigJson", "")
                if not cj_str:
                    continue
                try:
                    cj = json.loads(cj_str)
                except Exception:
                    continue
                nid = cj.get("ID")
                if isinstance(nid, int) and nid > 0:
                    used.add(nid)

    _GLOBAL_USED_IDS_CACHE[exclude_skill_id] = used
    return used


# ------------------------------------------------------------
# ID 分配器
# ------------------------------------------------------------
class IdAllocator:
    """分配 GUID + 自增 ConfigID。

    分配策略：
      - GUID: uuid4
      - SkillEffectConfig.ID: 从 32000000 + (skill_id % 1000) × 1000 起递增
      - SkillSelectConfig.ID: 同段位 +500 起
      - SkillConditionConfig.ID: 同段位 +800 起
      - SkillTagsConfig.ID: 优先用 IR tags 声明的；新建用 320900+ 段

    v2.5 (PostMortem #021): 初始化时预扫全工程已用 ID（除自身），分配时跳过冲突。
      避免如「skill_id 后 3 位相同」必撞 effect ID 的旧问题。
    """

    def __init__(self, skill_id: int, skip_global_scan: bool = False):
        self.skill_id = skill_id
        # 简易段位分配
        base = (skill_id % 1000) * 1000
        self._effect_seq = 32000000 + base       # SkillEffectConfig
        self._select_seq = 32000000 + base + 500
        self._cond_seq   = 32000000 + base + 800
        self._tag_seq    = 320900
        # v2.5：预填全项目已用 ID（除自身），分配时 skip
        if skip_global_scan:
            self._used_ids: set[int] = set()
        else:
            self._used_ids = set(_scan_global_used_ids(exclude_skill_id=skill_id))

    def new_guid(self) -> str:
        return str(uuid.uuid4())

    def allocate_effect_id(self) -> int:
        while self._effect_seq in self._used_ids:
            self._effect_seq += 1
        v = self._effect_seq
        self._used_ids.add(v)
        self._effect_seq += 1
        return v

    def allocate_select_id(self) -> int:
        while self._select_seq in self._used_ids:
            self._select_seq += 1
        v = self._select_seq
        self._used_ids.add(v)
        self._select_seq += 1
        return v

    def allocate_cond_id(self) -> int:
        while self._cond_seq in self._used_ids:
            self._cond_seq += 1
        v = self._cond_seq
        self._used_ids.add(v)
        self._cond_seq += 1
        return v

    def allocate_tag_id(self) -> int:
        while self._tag_seq in self._used_ids:
            self._tag_seq += 1
        v = self._tag_seq
        self._used_ids.add(v)
        self._tag_seq += 1
        return v

    def reserve(self, *ids: int):
        for i in ids:
            self._used_ids.add(i)


# ------------------------------------------------------------
# 模板调用节点生成（TSET_RUN_SKILL_EFFECT_TEMPLATE）
# ------------------------------------------------------------
def make_template_call_node(
    template_key: str,
    extra_params: list,
    alloc: "IdAllocator",
    ctx: "BuildCtx",
    desc: str = "",
) -> "Node":
    """生成 TSET_RUN_SKILL_EFFECT_TEMPLATE 节点。

    Params 顺序固定：
      [0] 主体单位 = entity:主体
      [1] 目标筛选 = entity:目标
      [2] 模板根 SkillEffectID（KNOWN_TEMPLATES 中的 root_effect_id）
      [3..] 用户提供的 extra_params（按模板的 TemplateParams 顺序）

    extra_params 的每项可以是：int / TParam / 引用字符串（"entity:主体" 等）。
    """
    tpl = KNOWN_TEMPLATES.get(template_key)
    if tpl is None:
        raise CompileError(f"未知模板: {template_key}（请在 KNOWN_TEMPLATES 中注册）")

    template_path = tpl["path"]
    root_id = tpl["root_effect_id"]

    # 拼接 Params
    params: list[TParam] = [
        resolve_ref("entity:主体"),
        resolve_ref("entity:目标"),
        TParam(value=root_id),
    ]
    for ep in extra_params:
        params.append(resolve_ref(ep) if not isinstance(ep, TParam) else ep)

    # TemplateData：从模板 JSON 加载 TemplateParams 复制过来
    template_params_def = load_template_params(template_path)

    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_RUN_SKILL_EFFECT_TEMPLATE",
        config_id=alloc.allocate_effect_id(),
        desc=desc or f"调用模板:{template_key}",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_RUN_SKILL_EFFECT_TEMPLATE"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_RUN_SKILL_EFFECT_TEMPLATE"),
            "TemplateData": {
                "TemplateParams": template_params_def,
                "TemplatePath": template_path,
            },
        },
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


# ------------------------------------------------------------
# 引用节点生成：BulletConfigNode / ModelConfigNode
# ------------------------------------------------------------
def make_bullet_config_node(
    bullet_id: int,
    alloc: IdAllocator,
    ctx: "BuildCtx",
    after_born_effect_id: int = 0,
) -> Node:
    """生成 BulletConfigNode：ConfigJson inline BulletConfig 表对应 ID 的全部字段。

    若读不到表（xlsx 不可用 / ID 不存在），降级填只含 ID 的最小字段。

    after_born_effect_id：子弹出生后执行的 SkillEffect ID（通常是碰撞模板节点）。
    设到 AfterBornSkillEffectExecuteInfo.SkillEffectConfigID 字段。
    """
    # 优先从已有 SkillGraph JSON 中拷贝合法节点（字段最完整、类型最正确）
    known = get_known_ref_config("BulletConfigNode", bullet_id)
    if known is not None:
        config_payload = dict(known)
        ctx.info(f"BulletConfig {bullet_id} 从已知合法节点拷贝（{len(config_payload)} 字段）")
    else:
        # 降级：从 xlsx 读取并过滤
        row = load_table_row("BulletConfig", bullet_id)
        if row is None:
            ctx.info(f"WARN: BulletConfig 表中找不到 ID={bullet_id}，降级生成空壳节点")
            config_payload = {"ID": bullet_id}
        else:
            config_payload = _filter_safe_fields(row)
            config_payload["ID"] = bullet_id

    # 关键：设定 AfterBornSkillEffectExecuteInfo（指向碰撞模板节点）
    if after_born_effect_id:
        config_payload["AfterBornSkillEffectExecuteInfo"] = {
            "SelectConfigID": 0,
            "SkillEffectConfigID": after_born_effect_id,
        }

    return Node(
        guid=alloc.new_guid(),
        cls="BulletConfigNode",
        config_id=bullet_id,
        desc="子弹属性",
        config_payload=config_payload,
        table_name="BulletConfig",
    )


def make_model_config_node(model_id: int, alloc: IdAllocator, ctx: "BuildCtx",
                            desc: str = "") -> Node:
    """生成 ModelConfigNode：inline ModelConfig 表数据。"""
    known = get_known_ref_config("ModelConfigNode", model_id)
    if known is not None:
        config_payload = dict(known)
        ctx.info(f"ModelConfig {model_id} 从已知合法节点拷贝（{len(config_payload)} 字段）")
    else:
        row = load_table_row("ModelConfig", model_id)
        if row is None:
            ctx.info(f"WARN: ModelConfig 表中找不到 ID={model_id}，降级生成空壳节点")
            config_payload = {"ID": model_id}
        else:
            config_payload = _filter_safe_fields(row)
            config_payload["ID"] = model_id

    return Node(
        guid=alloc.new_guid(),
        cls="ModelConfigNode",
        config_id=model_id,
        desc=desc or config_payload.get("ModelPath", ""),
        config_payload=config_payload,
        table_name="ModelConfig",
    )


# ------------------------------------------------------------
# Pattern Expander：直线子弹
# ------------------------------------------------------------
def _make_get_caster_facing_node(alloc: IdAllocator) -> Node:
    """生成 TSET_GET_ENTITY_ATTR_VALUE 节点读取施法者的 attr=91 (面向)。
    用于 angles 多颗子弹共享 facing 计算。
    """
    params = [
        TParam(value=75, param_type=1),   # PT_ATTR=1, 75=实体实例ID(取施法者)
        TParam(value=91),                 # 91=面向
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_GET_ENTITY_ATTR_VALUE",
        config_id=alloc.allocate_effect_id(),
        desc="读取施法者面向",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_GET_ENTITY_ATTR_VALUE") or 32,
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_GET_ENTITY_ATTR_VALUE") or 32},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


def _make_facing_offset_calc_node(facing_eid: int, offset_deg: int, alloc: IdAllocator) -> Node:
    """生成 TSET_NUM_CALCULATE 节点：facing + offset_deg。"""
    OP_ADD = 3
    params = [
        TParam(value=facing_eid, param_type=2),  # PT_FUNC_RET = 2
        TParam(value=OP_ADD),
        TParam(value=offset_deg),
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_NUM_CALCULATE",
        config_id=alloc.allocate_effect_id(),
        desc=f"facing + {offset_deg}°",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_NUM_CALCULATE") or 31,
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_NUM_CALCULATE") or 31},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


def _make_quick_damage_node(explosion_cfg: dict, alloc: IdAllocator) -> Node:
    """[DEPRECATED in v1.1.1] explosion → TSET_QUICK_DAMAGE。

    ⚠️ 此函数在 v1.1.1 之后已不再被 expand_bullet_straight 调用。
    保留代码仅供其他 builder 直接需要"对命中目标额外造一笔固定伤害"时使用。

    设计修正历史：
    v1.1 设计错误：用此节点实现"AOE 范围伤害"——导致一次命中造双份伤害（直伤模板 + QUICK_DAMAGE）
    v1.1.1 修正：explosion 改为语义糖，radius→collision_radius，damage→覆盖直伤参数
                 子弹通用-碰撞模板的 collision_radius 已是 AOE 检测半径，无需画蛇添足
    详见 PostMortem 2026-05-08-003 / docs/易错点速查.md §11

    Params 布局（按 SkillEffectConfig_ParamsAutoGenerated.cs §快速伤害结算）：
      [0] 伤害值（直接固定数值，**v1.1 限制**：不是 ATK× 公式）
      [1] 五行类型 TElementsType
      [2] 是否暴击 (bool)
      [3] 禁止伤害冒字 (bool)
      [4] 是否化解 (bool)
      [5] 伤害类型 TSkillDamageType
      [6] 是否闪避 (bool)
      [7] 死亡击飞配置 ID
      [8] 是否无视护盾 (bool)
      [9] 子类型 flags

    explosion_cfg.damage = {coef, type, element, subtype_flags}
    explosion_cfg.radius / target_camp 在 v1.1 暂不参与节点参数（v1.2 接 SkillSelect 圆形筛选时再用）
    """
    dmg = explosion_cfg.get("damage") or {}
    elem_map = {"随技能": 0, "金": 1, "木": 2, "水": 3, "火": 4, "土": 5}
    type_map = {"随技能": 0, "物理": 1, "法术": 2, "真实": 3}
    elem_v = dmg.get("element", "随技能")
    elem_int = elem_map.get(elem_v, 0) if isinstance(elem_v, str) else int(elem_v)
    type_v = dmg.get("type", "随技能")
    type_int = type_map.get(type_v, 0) if isinstance(type_v, str) else int(type_v)

    params = [
        TParam(value=dmg.get("coef", 0)),                       # [0] 伤害值
        TParam(value=elem_int),                                  # [1] 五行
        TParam(value=0),                                         # [2] 暴击=否
        TParam(value=0),                                         # [3] 禁冒字=否
        TParam(value=0),                                         # [4] 化解=否
        TParam(value=type_int),                                  # [5] 伤害类型
        TParam(value=0),                                         # [6] 闪避=否
        TParam(value=0),                                         # [7] 击飞=无
        TParam(value=0),                                         # [8] 无视护盾=否
        TParam(value=dmg.get("subtype_flags", 16)),              # [9] 子类型 flags
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_QUICK_DAMAGE",
        config_id=alloc.allocate_effect_id(),
        desc=f"爆炸伤害 (v1.1 单目标量 R={explosion_cfg.get('radius', 0)})",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_QUICK_DAMAGE") or 6,
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_QUICK_DAMAGE") or 6},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


def _make_on_hit_buff_node(buff_cfg: dict, alloc: IdAllocator) -> Node:
    """v1.1 on_hit_buff → TSET_ADD_BUFF（命中目标挂 buff）。复用 expand_apply_buff 的 Params 布局。"""
    params = [
        resolve_ref(buff_cfg.get("target", "entity:目标")),
        resolve_ref("entity:主体"),                       # source
        TParam(value=buff_cfg["buff_id"]),
        TParam(value=buff_cfg.get("duration_frames", 0)),
        TParam(value=0),                                  # 间隔帧（buff 内部决定）
        TParam(value=buff_cfg.get("stack", 1)),
        TParam(value=1),                                  # rapid_affect
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_ADD_BUFF",
        config_id=alloc.allocate_effect_id(),
        desc=f"命中挂 Buff #{buff_cfg['buff_id']}",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ADD_BUFF"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ADD_BUFF")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


def expand_bullet_straight(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """直线子弹模式 → TSET_CREATE_BULLET + BulletConfigNode + 命中链路（碰撞/伤害/表现 模板）。

    v1.1 新增：
      - cfg.angles: list[int] 多颗子弹角度偏移（基于 caster.facing），N 颗共享 BulletConfig 和命中链路
      - cfg.hit.explosion: 命中后造爆炸量伤害（TSET_QUICK_DAMAGE，v1.1 单目标量；v1.2 接 AOE 筛选）
      - cfg.hit.on_hit_buff: 命中后挂 buff（TSET_ADD_BUFF）

    命中链路结构（命中后 ORDER 顺序：直伤 → explosion → on_hit_buff → 表现）：
      BulletConfigNode.AfterBornSkillEffectExecuteInfo
            ↓
      子弹通用碰撞模板
            └ Params[3]=命中后功能 → 命中 ORDER
                                    ├ 子弹通用伤害模板
                                    ├ TSET_QUICK_DAMAGE (explosion，可选)
                                    ├ TSET_ADD_BUFF (on_hit_buff，可选)
                                    └ 子弹通用表现模板
    """
    cfg = step["bullet"]
    bullet_id = cfg["bullet_id"]
    creator   = resolve_ref(cfg.get("creator", "entity:主体"))
    offset_forward = cfg.get("offset_forward", 0)
    offset_right   = cfg.get("offset_right", 0)
    z_height       = cfg.get("z_height", 0)
    is_projectile  = 1 if cfg.get("is_projectile", True) else 0
    rapid_affect   = 1 if cfg.get("rapid_affect", False) else 0
    desc           = cfg.get("desc", "创建子弹")
    angles         = cfg.get("angles") or []   # v1.1: 多颗子弹角度偏移列表

    # 多颗模式下，预生成共享的 caster.facing 节点（单颗或 angles 全 0 时不需要）
    facing_node: Node | None = None
    if angles and any(a != 0 for a in angles):
        facing_node = _make_get_caster_facing_node(alloc)

    def _build_create_bullet(angle_offset: int, calc_node: Node | None) -> Node:
        """生成单颗 TSET_CREATE_BULLET 节点。calc_node 非 None 时表示 facing 用 FUNC_RET。"""
        if calc_node is None:
            facing_param = resolve_ref("attr:面向")
        else:
            facing_param = TParam(value=calc_node.config_id, param_type=2)
        params_local = [
            TParam(value=bullet_id),
            facing_param,
            resolve_ref("attr:位置X"),
            resolve_ref("attr:位置Y"),
            creator,
            TParam(value=offset_right),
            TParam(value=offset_forward),
            TParam(value=0),
            TParam(value=0),
            TParam(value=is_projectile),
            TParam(value=0),
            TParam(value=0),
            TParam(value=z_height),
            TParam(value=0),
            TParam(value=rapid_affect),
        ]
        bn = Node(
            guid=alloc.new_guid(),
            cls="TSET_CREATE_BULLET",
            config_id=alloc.allocate_effect_id(),
            desc=f"{desc} (offset={angle_offset}°)" if angles else desc,
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_BULLET"),
                "Params": [p.to_dict() for p in params_local],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_BULLET")},
            table_name="SkillEffectConfig",
        )
        bn.config_payload["ID"] = bn.config_id
        return bn

    # 生成 N 颗 CREATE_BULLET（angles 为空时退化为 1 颗，与 v1.0 行为一致）
    bullet_nodes: list[Node] = []
    facing_calc_nodes: list[Node] = []
    if not angles:
        bullet_nodes.append(_build_create_bullet(0, None))
    else:
        for offset in angles:
            calc_node = None
            if offset != 0:
                calc_node = _make_facing_offset_calc_node(facing_node.config_id, offset, alloc)
                facing_calc_nodes.append(calc_node)
            bullet_nodes.append(_build_create_bullet(offset, calc_node))

    # 多颗时用 ORDER 包起来
    fan_order_node: Node | None = None
    if len(bullet_nodes) > 1:
        params_order = [TParam(value=bn.config_id) for bn in bullet_nodes]
        fan_order_node = Node(
            guid=alloc.new_guid(),
            cls="TSET_ORDER_EXECUTE",
            config_id=alloc.allocate_effect_id(),
            desc=f"扇形发射 {len(bullet_nodes)} 颗（angles={angles}）",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
                "Params": [p.to_dict() for p in params_order],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
            table_name="SkillEffectConfig",
        )
        fan_order_node.config_payload["ID"] = fan_order_node.config_id

    # 用第一颗作为后续 hit 链路构建的"代表"（因为命中链路与具体哪一颗无关，由 BulletConfig.AfterBorn 引用）
    create_bullet_node = bullet_nodes[0]

    # 命中链路（按 IR 中的 hit 块构建）
    extra_nodes: list[Node] = []
    hit_cfg = cfg.get("hit") or {}
    after_born_id = 0  # BulletConfigNode.AfterBornSkillEffectExecuteInfo 指向

    if hit_cfg:
        # === 1) 命中表现模板 ===
        hit_effect_model = hit_cfg.get("hit_effect_model", 0)
        present_node = make_template_call_node(
            template_key="子弹通用逻辑-表现",
            extra_params=[
                0,                          # 自定义额外效果
                0,                          # 基础模式 0=命中目标才播放
                hit_effect_model,           # 命中特效资源 ModelConfig
                2,                          # 命中特效位置 = 2 默认
                100,                        # 缩放%
                2,                          # 喷射强度 = 2 正常
                0,                          # 顿帧
                0, 0, 0,                    # 受击震屏
                1,                          # 主角参与命中顿帧
                0, 0, 0,                    # 受击动作 / 击退表现 / 受击音效
                0, 0, 0, False, 30, 0,      # 屏幕模糊 / 波纹 / 闪红
            ],
            alloc=alloc, ctx=ctx,
            desc="子弹通用-表现",
        )
        extra_nodes.append(present_node)

        # 命中表现节点的额外 ModelConfigNode（如果指定了 hit_effect_model）
        if hit_effect_model:
            extra_nodes.append(make_model_config_node(
                hit_effect_model, alloc, ctx, desc="命中特效",
            ))

        # === 2) 命中伤害模板 ===
        # v1.1 设计修正（PostMortem 2026-05-08-003）：
        #   原方案误以为需要 explosion → TSET_QUICK_DAMAGE 实现"圆形 AOE 多目标"。
        #   实际上：子弹通用-碰撞模板的 collision_radius 已经是 AOE 检测半径，
        #          范围内每个目标都会各调用一次伤害模板（即天然 AOE 多目标伤害）。
        #   所以正确语义：
        #     - hit.damage          → 直伤模板的伤害参数（默认）
        #     - hit.explosion       → 表示"我要 AOE 爆炸效果"：把 collision_radius 设为 explosion.radius
        #                             如果同时给了 explosion.damage，**用 explosion.damage 覆盖直伤参数**
        #                             （语义：爆炸量取代直伤量；不再生成额外的 QUICK_DAMAGE 节点）
        #
        # 决定 effective damage：explosion.damage 优先于 hit.damage
        explosion_cfg = hit_cfg.get("explosion") or {}
        damage_cfg = explosion_cfg.get("damage") or hit_cfg.get("damage") or {}

        damage_element_map = {"随技能": 0, "金": 1, "木": 2, "水": 3, "火": 4, "土": 5}
        damage_element = damage_cfg.get("element", "随技能")
        damage_element_int = damage_element_map.get(damage_element, 0) if isinstance(damage_element, str) else int(damage_element)
        damage_type_map = {"随技能": 0, "物理": 1, "法术": 2, "真实": 3}
        damage_type = damage_cfg.get("type", "随技能")
        damage_type_int = damage_type_map.get(damage_type, 0) if isinstance(damage_type, str) else int(damage_type)
        damage_coef = damage_cfg.get("coef", 16)
        damage_extra = damage_cfg.get("extra", -1)
        damage_subtype_flags = damage_cfg.get("subtype_flags", 16)

        damage_desc = "子弹通用-伤害（爆炸量）" if explosion_cfg else "子弹通用-伤害"
        damage_node = make_template_call_node(
            template_key="子弹通用逻辑-伤害",
            extra_params=[
                1,
                "entity:施法者伤害归属",
                damage_type_int,
                damage_element_int,
                damage_subtype_flags,
                damage_coef,
                damage_extra,
                0,
                0, 0,
                1 if hit_cfg.get("destroy_on_hit", True) else 0,
                -1,
                -1,
            ],
            alloc=alloc, ctx=ctx,
            desc=damage_desc,
        )
        extra_nodes.append(damage_node)

        # === 2.6) v1.1 新增：命中挂 buff（TSET_ADD_BUFF） ===
        on_hit_buff_node: Node | None = None
        if hit_cfg.get("on_hit_buff"):
            on_hit_buff_node = _make_on_hit_buff_node(hit_cfg["on_hit_buff"], alloc)
            extra_nodes.append(on_hit_buff_node)

        # === 3) 命中 ORDER（按顺序：伤害 → 挂 buff → 表现）===
        hit_order_children_ids: list[int] = [damage_node.config_id]
        if on_hit_buff_node:
            hit_order_children_ids.append(on_hit_buff_node.config_id)
        hit_order_children_ids.append(present_node.config_id)

        hit_order_node = Node(
            guid=alloc.new_guid(),
            cls="TSET_ORDER_EXECUTE",
            config_id=alloc.allocate_effect_id(),
            desc="命中后执行",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
                "Params": [TParam(value=cid).to_dict() for cid in hit_order_children_ids],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
            table_name="SkillEffectConfig",
        )
        hit_order_node.config_payload["ID"] = hit_order_node.config_id
        extra_nodes.append(hit_order_node)

        # === 4) 子弹碰撞模板 ===
        # v1.1 设计修正：explosion.radius 直接映射到 collision_radius（实现 AOE 检测）
        # 优先级：explosion.radius > hit.collision_radius > 默认 400
        collision_radius = (
            (hit_cfg.get("explosion") or {}).get("radius")
            or hit_cfg.get("collision_radius", 400)
        )
        collision_height = hit_cfg.get("collision_height", 360)
        # 范围类型: 1=圆形 / 2=矩形 / 3=扇形 / 4=环形
        shape_type = 1
        target_camp = 0  # 0=敌军 / 1=所有 / 2=友军 / 3=自己 / 4=友军+自己
        # PostMortem #036: 侦测间隔默认 1 / 侦测冷却默认 10 (与碰撞模板自身默认一致 + 用户实战经验值)
        # 旧默认 detect_interval=10/侦测冷却=0 是反向值 → 已修订
        detect_interval = hit_cfg.get("detect_interval", 1)
        detect_cooldown = hit_cfg.get("detect_cooldown", 10)

        collision_node = make_template_call_node(
            template_key="子弹通用逻辑-碰撞",
            extra_params=[
                hit_order_node.config_id,            # 命中后功能
                shape_type,                          # 碰撞范围类型
                collision_radius,                    # 参数1（圆形:半径）
                collision_height,                    # 参数2
                0,                                   # 参数3
                offset_right,                        # 位置偏移右
                offset_forward,                      # 位置偏移前
                target_camp,                         # 目标单位类型
                0,                                   # 自定义条件
                1,                                   # 是否跟随创建者
                detect_interval,                     # 侦测间隔 (PostMortem #036 默认 1)
                detect_cooldown,                     # 侦测冷却 (PostMortem #036 默认 10)
                0,                                   # 侦测次数
            ],
            alloc=alloc, ctx=ctx,
            desc="子弹通用-碰撞",
        )
        extra_nodes.append(collision_node)
        after_born_id = collision_node.config_id

    # 子弹本体节点（带 AfterBorn 指向碰撞模板）— N 颗子弹共享同一个 BulletConfigNode
    bullet_cfg_node = make_bullet_config_node(bullet_id, alloc, ctx, after_born_effect_id=after_born_id)

    # v1.1 多颗模式：返回 ORDER 作为代表节点 + 所有 N 颗 + facing 节点 + facing offset CALC 节点 + 命中链路 + BulletConfig
    if fan_order_node:
        # 代表节点是 fan_order_node（被父级 delay/order 引用）
        all_nodes = [fan_order_node] + bullet_nodes
        if facing_node:
            all_nodes.append(facing_node)
        all_nodes.extend(facing_calc_nodes)
        all_nodes.append(bullet_cfg_node)
        all_nodes.extend(extra_nodes)
        return all_nodes

    # 单颗模式（v1.0 行为完全一致）
    return [create_bullet_node, bullet_cfg_node] + extra_nodes


# ------------------------------------------------------------
# Pattern Expander：回旋飞回（v1.2 新增）
# ------------------------------------------------------------
def _make_inline_bullet_config_node(
    bullet_id: int,
    fly_type: int,
    trace_path_type: int,
    trace_path_params: list,
    speed: int,
    last_time: int,
    max_distance: int,
    after_born_id: int,
    die_skill_id: int,
    alloc: IdAllocator,
    ctx: "BuildCtx",
    desc: str = "",
) -> Node:
    """v1.2 内联生成定制 BulletConfigNode（不依赖已有 BulletConfig 表行）。

    用于回旋飞回 pattern：飞出阶段 Bullet1 + 回飞阶段 Bullet2 都用此函数构造。
    """
    # 优先从已知合法节点拷贝基础字段（保证 ConfigJson 字段完整）
    known = get_known_ref_config("BulletConfigNode", 320146)  # 复用 30212005 木影旋刃的字段模板
    if known is not None:
        config_payload = dict(known)
    else:
        # 降级：用 320032 的（直线子弹模板）
        known2 = get_known_ref_config("BulletConfigNode", 320032)
        config_payload = dict(known2) if known2 is not None else {}

    # 用 v1.2 参数覆盖关键字段
    config_payload["ID"] = bullet_id
    config_payload["FlyType"] = fly_type
    config_payload["TracePathType"] = trace_path_type
    if trace_path_params:
        config_payload["TracePathParams"] = [
            {"Value": v, "ParamType": 0, "Factor": 0} for v in trace_path_params
        ]
    if speed > 0:
        config_payload["Speed"] = speed
    if last_time > 0:
        config_payload["LastTime"] = last_time
    if max_distance > 0:
        config_payload["MaxDistance"] = max_distance

    # 三个生命周期钩子
    config_payload["BeforeBornSkillEffectExecuteInfo"] = {"SelectConfigID": 0, "SkillEffectConfigID": 0}
    config_payload["AfterBornSkillEffectExecuteInfo"] = {
        "SelectConfigID": 0, "SkillEffectConfigID": after_born_id
    }
    config_payload["DieSkillEffectExecuteInfo"] = {
        "SelectConfigID": 0, "SkillEffectConfigID": die_skill_id
    }

    return Node(
        guid=alloc.new_guid(),
        cls="BulletConfigNode",
        config_id=bullet_id,
        desc=desc or f"Bullet({bullet_id}) FlyType={fly_type}",
        config_payload=config_payload,
        table_name="BulletConfig",
    )


# ------------------------------------------------------------
# Pattern Expander：链状指示器（v2.5 新增 / 端绑型 Beam）
# ------------------------------------------------------------
# 设计基底（280103 狮妖连线金标）：
#   BulletConfig.FlyType=0(静态) + Model=4(空锚点) + ChainModel=2800355(链视觉)
#   BulletConfig.BeforeBornSE = TSET_FOLLOW_ENTITY (Type=30)
#       子弹本体跟随"最初施法者"（玩家主角）
#       Params: [V=75 PT=1(attr,子弹自身)] [V=35 PT=5(施法者-根创建者)] [0,0] [V=1 PT=0 位置同步=开]
#       ⚠️ 反直觉点（r2 修订 2026-05-13 用户裁决）：
#         BulletConfig.BeforeBornSE / AfterBornSE / DieSE 是嵌套 SkillEffect 调用栈，
#         此时 V=1 TCPT_MAIN_ENTITY "主体" = 子弹自身（不是玩家主角！）。
#         要拿"最初施法的玩家主角"必须用 V=35 TCPT_CREATE_ROOT_CREATOR_ENTITY_ID
#         （施法者-根创建者；280103 金标用 V=37 主体创建者，V=35 更鲁棒避免被
#          buff/嵌套技能链覆盖）。
#         同源陷阱波及 Buff.OnAttachEffect / OnTickEffect 等所有嵌套 SkillEffect 上下文。
#   BulletConfig.AfterBornSE = TSET_MODIFY_ENTITY_ATTR_VALUE (Type=12)
#       把子弹的 TBN_CHAIN_TARGET_ENTITY(=118) 改为目标实体 → 链端绑目标
#       Params: [V=75 PT=1(attr,子弹自身)] [V=118 PT=0(改的 attr 编号)] [V=2 PT=5(目标单位实例ID)]
#   CREATE_BULLET 把 BulletConfig 实例化在主体位置 / 不投射(is_projectile=0)
#       Params[4] 创建者 = V=35 施法者-根创建者 (与 BeforeBornSE 一致；金标 V=37 同精神)
#
# 与 280103 一一对应（仅角色对调：280103 主体=怪物 → 本任务主体=玩家主角）

def _make_chain_follow_entity_node(target_param: TParam, alloc: IdAllocator,
                                    desc: str = "子弹端绑施法者根创建者(玩家)") -> Node:
    """TSET_FOLLOW_ENTITY (Type=30) 节点 — 子弹本体跟随某实体。

    Params:
      [0] 跟随主体 = {V=75, PT=1}(attr,子弹自身实例ID)
      [1] 被跟随目标 = target_param（r2 修订默认 entity:施法者根 → {V=35, PT=5}）
            ⚠️ 此节点位于 BulletConfig.BeforeBornSE 嵌套上下文 / V=1 "主体"=子弹自身 /
               要拿玩家主角必须 V=35 TCPT_CREATE_ROOT_CREATOR_ENTITY_ID
      [2] 0
      [3] 位置同步 = 1
      [4..7] 0
    """
    params = [
        TParam(value=75, param_type=1),  # 子弹自身
        target_param,                     # 被跟随的目标实体
        TParam(value=0),
        TParam(value=1),                  # 位置同步
        TParam(value=0), TParam(value=0), TParam(value=0), TParam(value=0),
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_FOLLOW_ENTITY",
        config_id=alloc.allocate_effect_id(),
        desc=desc,
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_FOLLOW_ENTITY") or 30,
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_FOLLOW_ENTITY") or 30},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


def _make_chain_bind_target_node(target_param: TParam, alloc: IdAllocator,
                                   desc: str = "链端绑目标") -> Node:
    """TSET_MODIFY_ENTITY_ATTR_VALUE (Type=12) 节点 — 把子弹的链目标 attr 设为某实体。

    TBattleNatureEnum.TBN_CHAIN_TARGET_ENTITY = 118 (见 skill_editor_enums.json)
    底层走 BattleBulletAttrCollectComp → effectChainComp.SetTarget(newValue)

    Params:
      [0] 修改实体 = {V=75, PT=1}(attr,子弹自身实例ID)
      [1] 改的 attr 编号 = {V=118, PT=0}
      [2] 新值 = target_param（默认 entity:目标 → {V=2, PT=5}）
    """
    params = [
        TParam(value=75, param_type=1),   # 子弹自身
        TParam(value=118, param_type=0),  # TBN_CHAIN_TARGET_ENTITY
        target_param,                      # 链目标实体
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_MODIFY_ENTITY_ATTR_VALUE",
        config_id=alloc.allocate_effect_id(),
        desc=desc,
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_MODIFY_ENTITY_ATTR_VALUE") or 12,
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_MODIFY_ENTITY_ATTR_VALUE") or 12},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


def _make_chain_bullet_config_node(
    bullet_id: int,
    chain_model_id: int,
    life_frames: int,
    before_born_id: int,
    after_born_id: int,
    alloc: IdAllocator,
    ctx: "BuildCtx",
    desc: str = "",
) -> Node:
    """v2.5 链状指示器专用 BulletConfigNode（FlyType=0 静态 + Model=4 空锚点 + ChainModel）。

    基底字段：复用 280103 金标 BulletConfig=280211（5591 行 JSON 实测）。
    若拿不到 280211（理论上扫缓存能拿到），降级用最小字段集。

    关键字段（与 280103 280211 1:1 对应）：
      ID=bullet_id（IP=250 段）
      FlyType=0（静态 / 不投射）
      Model=4（空锚点 / PostMortem 创建子弹必须 Model=4）
      ChainModel=chain_model_id（链视觉模型 / 默认 2800355）
      ChainModelScalePercent=100 / ChainTilingFactor=100
      LastTime=life_frames（命中段时长 / 30帧=1秒）
      DestroyWhenCreatorDie=true
      LifeFlag=1
      BeforeBornSkillEffectExecuteInfo.SkillEffectConfigID = before_born_id (FOLLOW_ENTITY)
      AfterBornSkillEffectExecuteInfo.SkillEffectConfigID  = after_born_id  (MODIFY_ATTR 链端绑目标)
      DieSkillEffectExecuteInfo.SkillEffectConfigID = 0
    """
    # 优先从已知 280211 复用字段模板
    known = get_known_ref_config("BulletConfigNode", 280211)
    if known is not None:
        config_payload = dict(known)
    else:
        config_payload = {}

    config_payload["ID"] = bullet_id
    config_payload["FlyType"] = 0
    config_payload["Model"] = 4
    config_payload["ChainModel"] = chain_model_id
    config_payload["ChainModelScalePercent"] = 100
    config_payload["ChainTilingFactor"] = 100
    config_payload["LastTime"] = life_frames
    config_payload["DestroyWhenCreatorDie"] = True
    config_payload["LifeFlag"] = 1
    # 清掉 280103 残留可能不适用的字段
    config_payload["Speed"] = 0
    config_payload["MaxDistance"] = 0
    config_payload["TracePathType"] = 0
    config_payload["TracePathParams"] = []
    config_payload["delayDestroyTime"] = 0
    config_payload["BeforeBornSkillEffectExecuteInfo"] = {
        "SelectConfigID": 0,
        "SkillEffectConfigID": before_born_id,
    }
    config_payload["AfterBornSkillEffectExecuteInfo"] = {
        "SelectConfigID": 0,
        "SkillEffectConfigID": after_born_id,
    }
    config_payload["DieSkillEffectExecuteInfo"] = {
        "SelectConfigID": 0,
        "SkillEffectConfigID": 0,
    }

    ctx.info(
        f"链状指示器 BulletConfig {bullet_id}: FlyType=0 Model=4 "
        f"ChainModel={chain_model_id} LastTime={life_frames} "
        f"BeforeBorn={before_born_id} AfterBorn={after_born_id}"
    )

    return Node(
        guid=alloc.new_guid(),
        cls="BulletConfigNode",
        config_id=bullet_id,
        desc=desc or f"链子弹 ChainModel={chain_model_id}",
        config_payload=config_payload,
        table_name="BulletConfig",
    )


def _alloc_chain_bullet_id(scan_used_ids: set[int] | None = None) -> int:
    """从 IP=250 段（2500001 起步）分配一个未被全工程占用的 BulletConfig ID。"""
    if scan_used_ids is None:
        scan_used_ids = set()
        for fp in SKILL_JSON_ROOT.rglob("SkillGraph_*.json"):
            try:
                g = json.loads(fp.read_text(encoding="utf-8"))
            except Exception:
                continue
            for r in g.get("references", {}).get("RefIds", []):
                cls = r.get("type", {}).get("class", "")
                if not cls.endswith("BulletConfigNode"):
                    continue
                cj_str = r.get("data", {}).get("ConfigJson", "")
                if not cj_str:
                    continue
                try:
                    cj = json.loads(cj_str)
                except Exception:
                    continue
                v = cj.get("ID")
                if isinstance(v, int):
                    scan_used_ids.add(v)
    cur = 2500001
    while cur in scan_used_ids:
        cur += 1
    return cur


def expand_bullet_chain(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """v2.5 链状指示器模式。

    产出节点（5 个 + 1 SkillEffect ORDER 入口）：
      1. TSET_ORDER_EXECUTE (入口 / 单子)：→ [CREATE_BULLET]
      2. TSET_CREATE_BULLET：参数 P[0]=bullet_id, P[4]=主体, P[10]=41(初始技能实例ID)
      3. BulletConfigNode：FlyType=0 + Model=4 + ChainModel + BeforeBornSE + AfterBornSE
      4. TSET_FOLLOW_ENTITY (BeforeBorn)：子弹本体跟主体
      5. TSET_MODIFY_ENTITY_ATTR_VALUE (AfterBorn)：把子弹链端 attr=118 设为目标实体

    返回的第一个节点是 ORDER 入口（让上层 build_order 当父引用）。
    """
    cfg = step["bullet"]

    # 配置参数
    bullet_id     = cfg.get("bullet_id") or _alloc_chain_bullet_id()
    chain_model_id = cfg.get("chain_model_id", 2800355)  # 默认复用 280103 狮妖连线
    life_frames    = cfg.get("life_frames", 60)           # 默认 60 帧 = 2 秒（30 帧/秒铁律）
    bind_to        = cfg.get("bind_to", "目标")
    no_hit         = cfg.get("no_hit", True)
    desc           = cfg.get("desc", "链状指示器子弹")

    # ⚠️ creator / bind_master 默认 = "施法者根" (V=35) 而非 "主体" (V=1)
    #   r2 修订 2026-05-13：BulletConfig.BeforeBornSE 嵌套上下文中 V=1 主体=子弹自身，
    #   要拿玩家主角必须 V=35 施法者-根创建者。详见本文件顶部 Pattern Expander 注释。
    creator     = resolve_ref(cfg.get("creator", "entity:施法者根"))
    bind_master = resolve_ref(cfg.get("bind_master", "entity:施法者根"))
    offset_forward = cfg.get("offset_forward", 0)
    offset_right   = cfg.get("offset_right", 0)
    z_height       = cfg.get("z_height", 0)
    # 链状指示器子弹是"静态吸附" — 不投射 / is_projectile=0
    is_projectile  = 0
    rapid_affect   = 1 if cfg.get("rapid_affect", False) else 0

    # 目标端引用（默认 entity:目标）
    target_ref_map = {"目标": "entity:目标", "主体": "entity:主体", "施法者": "entity:施法者", "施法者根": "entity:施法者根"}
    target_ref_str = target_ref_map.get(bind_to, "entity:目标")
    target_param = resolve_ref(target_ref_str)

    # 1) BeforeBornSE：FOLLOW_ENTITY 子弹本体跟 施法者-根创建者(玩家主角)
    #    ⚠️ 嵌套 BulletConfig.BeforeBornSE 上下文 V=1 主体=子弹自身，必须 V=35（r2 修订）
    follow_node = _make_chain_follow_entity_node(
        bind_master, alloc,
        desc="子弹本体跟随施法者根创建者(玩家主角) / V=35 / 嵌套BulletConfig.BeforeBornSE 上下文里 V=1主体=子弹自身 必须用 V=35"
    )

    # 2) AfterBornSE：MODIFY_ATTR 把子弹的 TBN_CHAIN_TARGET_ENTITY=118 attr 设为目标
    bind_target_node = _make_chain_bind_target_node(
        target_param, alloc, desc=f"链端绑{bind_to}"
    )

    # 3) BulletConfig（FlyType=0 + Model=4 + ChainModel + 两个 BornSE 钩子）
    bullet_cfg_node = _make_chain_bullet_config_node(
        bullet_id=bullet_id,
        chain_model_id=chain_model_id,
        life_frames=life_frames,
        before_born_id=follow_node.config_id,
        after_born_id=bind_target_node.config_id,
        alloc=alloc, ctx=ctx,
        desc=f"链子弹({bullet_id}) Chain={chain_model_id} {life_frames}帧",
    )

    # 4) CREATE_BULLET 节点（Params 15 个 / 与 280103 28023952 1:1 对账）
    params_create = [
        TParam(value=bullet_id),                     # [0] BulletConfig ID
        resolve_ref("attr:面向"),                    # [1] 朝向 = attr 91
        resolve_ref("attr:位置X"),                   # [2] 位置X = attr 59
        resolve_ref("attr:位置Y"),                   # [3] 位置Y = attr 60
        creator,                                      # [4] 创建者 = 主体
        TParam(value=offset_right),                  # [5] 偏移右
        TParam(value=offset_forward),                # [6] 偏移前
        TParam(value=0),                              # [7]
        TParam(value=0),                              # [8]
        TParam(value=is_projectile),                 # [9] 是否投射（链状=0 静态）
        TParam(value=41, param_type=5),              # [10] 初始攻击实体ID = 初始技能实例ID
        TParam(value=0),                              # [11]
        TParam(value=z_height),                      # [12] z 高度
        TParam(value=0),                              # [13]
        TParam(value=rapid_affect),                  # [14] rapid_affect
    ]
    create_bullet_node = Node(
        guid=alloc.new_guid(),
        cls="TSET_CREATE_BULLET",
        config_id=alloc.allocate_effect_id(),
        desc=desc,
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_BULLET") or 8,
            "Params": [p.to_dict() for p in params_create],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_BULLET") or 8},
        table_name="SkillEffectConfig",
    )
    create_bullet_node.config_payload["ID"] = create_bullet_node.config_id

    # 5) 入口 ORDER（即使只有 1 个子节点也包一下 / 与样本风格一致）
    order_node = Node(
        guid=alloc.new_guid(),
        cls="TSET_ORDER_EXECUTE",
        config_id=alloc.allocate_effect_id(),
        desc="主入口-创建链状指示器子弹",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
            "Params": [TParam(value=create_bullet_node.config_id).to_dict()],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
        table_name="SkillEffectConfig",
    )
    order_node.config_payload["ID"] = order_node.config_id

    if not no_hit:
        ctx.info("WARN: 链状指示器 no_hit=false 未实现 / 当前按纯视觉处理")

    return [order_node, create_bullet_node, bullet_cfg_node, follow_node, bind_target_node]


def expand_bullet_boomerang(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """v1.2 回旋飞回模式（v1.2.1 简化版 — 单 Bullet 套蝴蝶妖回旋镖配置）。

    设计修订（PostMortem #007）：
        v1.2.0 双 Bullet 接力方案在实测中失败：Bullet2 (FlyType=6 TRACK_TARGET_POINT)
        没有正确的"追施法者"配置，飞向远处而非追回主角。

    v1.2.1 简化方案：
        套用蝴蝶妖回旋镖 (BulletConfig 2200124) 已验证的 TracePathType=3 折返配置
        - FlyType=5 / TracePathType=3 自带"飞出 + 悬停 + 追施法者飞回"全流程
        - TracePathParams=[25, 8, 1000, 0, 1500, 0] (蝴蝶妖同款)
        - Speed=2000 / AcceSpeed=1500 / MaxSpeed=5000
        - LastTime / MaxDistance 由 boomerang 字段控制
        - AfterBorn = 子弹通用-碰撞模板（**全程挂碰撞** — 包括飞出阶段）

    已知局限（用户已知情接受）：
        - 飞出阶段碰到目标也造伤（不严格匹配 LOL 泰隆 R 的"出去无伤"）
        - 但视觉效果（一圈飞出+悬停+追施法者飞回）与原版一致
        - 未来想精确"出去无伤 / 回飞才有伤"需要扩 IR 加 SkillTag 时间条件控制
    """
    cfg = step["bullet"]
    bullet1_id = cfg.get("bullet_id", 0)
    bullet2_id = cfg.get("return_bullet_id", 0) or (bullet1_id + 1 if bullet1_id else 0)
    if not bullet1_id:
        # 自动分配（用 skill_id 段位 + 50/51 偏移）
        bullet1_id = (alloc.skill_id % 1000) * 1000 + 32000050
        bullet2_id = bullet1_id + 1
    if not bullet2_id:
        bullet2_id = bullet1_id + 1

    creator   = resolve_ref(cfg.get("creator", "entity:主体"))
    offset_forward = cfg.get("offset_forward", 0)
    offset_right   = cfg.get("offset_right", 0)
    z_height       = cfg.get("z_height", 0)
    rapid_affect   = 1 if cfg.get("rapid_affect", False) else 0
    desc           = cfg.get("desc", "回旋飞回子弹")
    angles         = cfg.get("angles") or []

    # 回旋参数
    boom = cfg.get("boomerang") or {}
    flight_frames = boom.get("flight_frames", 30)
    hover_frames  = boom.get("hover_frames", 20)
    max_distance  = boom.get("max_distance", 600)
    return_speed  = boom.get("return_speed", 2000)
    bullet1_last_time = flight_frames + hover_frames

    # === 1) 命中链路（v1.2.1 简化版：单 Bullet 全程挂碰撞）===
    extra_nodes: list[Node] = []
    bullet_after_born_id = 0
    hit_cfg = cfg.get("hit") or {}
    if hit_cfg:
        # 1a) 表现模板
        hit_effect_model = hit_cfg.get("hit_effect_model", 0)
        present_node = make_template_call_node(
            template_key="子弹通用逻辑-表现",
            extra_params=[0, 0, hit_effect_model, 2, 100, 2, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, False, 30, 0],
            alloc=alloc, ctx=ctx, desc="子弹通用-表现",
        )
        extra_nodes.append(present_node)
        if hit_effect_model:
            extra_nodes.append(make_model_config_node(hit_effect_model, alloc, ctx, desc="命中特效"))

        # 1b) 伤害模板
        damage_cfg = hit_cfg.get("damage") or {}
        elem_map = {"随技能": 0, "金": 1, "木": 2, "水": 3, "火": 4, "土": 5}
        type_map = {"随技能": 0, "物理": 1, "法术": 2, "真实": 3}
        d_elem = damage_cfg.get("element", "随技能")
        d_elem_int = elem_map.get(d_elem, 0) if isinstance(d_elem, str) else int(d_elem)
        d_type = damage_cfg.get("type", "随技能")
        d_type_int = type_map.get(d_type, 0) if isinstance(d_type, str) else int(d_type)
        damage_node = make_template_call_node(
            template_key="子弹通用逻辑-伤害",
            extra_params=[
                1, "entity:施法者伤害归属",
                d_type_int, d_elem_int,
                damage_cfg.get("subtype_flags", 16),
                damage_cfg.get("coef", 12000),
                damage_cfg.get("extra", -1),
                0, 0, 0,
                1 if hit_cfg.get("destroy_on_hit", True) else 0,
                -1, -1,
            ],
            alloc=alloc, ctx=ctx, desc="子弹通用-伤害",
        )
        extra_nodes.append(damage_node)

        # 1c) 命中后 ORDER
        hit_order_node = Node(
            guid=alloc.new_guid(),
            cls="TSET_ORDER_EXECUTE",
            config_id=alloc.allocate_effect_id(),
            desc="命中后执行",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
                "Params": [
                    TParam(value=damage_node.config_id).to_dict(),
                    TParam(value=present_node.config_id).to_dict(),
                ],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
            table_name="SkillEffectConfig",
        )
        hit_order_node.config_payload["ID"] = hit_order_node.config_id
        extra_nodes.append(hit_order_node)

        # 1d) 子弹通用-碰撞模板【第一个】检测敌方造伤害
        collision_radius = hit_cfg.get("collision_radius") or 100
        collision_height = hit_cfg.get("collision_height") or 200
        detect_interval = hit_cfg.get("detect_interval", 5)
        collision_enemy_node = make_template_call_node(
            template_key="子弹通用逻辑-碰撞",
            extra_params=[
                hit_order_node.config_id,    # [0] 命中后功能
                0,                            # [1] 碰撞范围类型 = 0 圆形 (TTriggerType)
                collision_radius,             # [2] 半径
                collision_height,             # [3] 参数2
                0,                            # [4] 参数3
                offset_right,                 # [5] 位置偏移右
                offset_forward,               # [6] 位置偏移前
                0,                            # [7] 目标单位类型 = 0 敌军
                0,                            # [8] 自定义条件
                0,                            # [9] 是否跟随创建者飞行
                detect_interval,              # [10] 侦测间隔
                0,                            # [11] 侦测冷却
                0,                            # [12] 侦测次数
            ],
            alloc=alloc, ctx=ctx, desc="子弹通用-碰撞[敌方]（造伤害）",
        )
        extra_nodes.append(collision_enemy_node)

        # === v1.3.1 DELAY 包第二碰撞模板（PostMortem #012 教训修正）===
        # 用 TSET_DELAY_EXECUTE 在飞出+悬停阶段过后才挂载主角碰撞检测
        # → 子弹出生时不立即销毁，回飞时才开始检测主角
        # 1e) 销毁子弹节点
        destroy_bullet_node = Node(
            guid=alloc.new_guid(),
            cls="TSET_DESTROY_ENTITY",
            config_id=alloc.allocate_effect_id(),
            desc="销毁子弹（碰到主角）",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_DESTROY_ENTITY") or 24,
                "Params": [
                    resolve_ref("entity:主体").to_dict(),
                    TParam(value=0).to_dict(),
                    TParam(value=0).to_dict(),
                ],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_DESTROY_ENTITY") or 24},
            table_name="SkillEffectConfig",
        )
        destroy_bullet_node.config_payload["ID"] = destroy_bullet_node.config_id
        extra_nodes.append(destroy_bullet_node)

        # 1f) 第二碰撞模板：检测主角（target_camp=3）→ 销毁子弹
        return_collision_radius = 100
        collision_self_node = make_template_call_node(
            template_key="子弹通用逻辑-碰撞",
            extra_params=[
                destroy_bullet_node.config_id,  # [0] 命中后功能 = 销毁子弹
                0,                              # [1] 圆形
                return_collision_radius,        # [2] 半径
                collision_height,               # [3]
                0,                              # [4]
                0,                              # [5]
                0,                              # [6]
                3,                              # [7] 目标单位类型 = 3 自己（创建者）
                0,                              # [8]
                0,                              # [9] 不跟随创建者
                5,                              # [10] 侦测间隔
                0,                              # [11] 侦测冷却
                0,                              # [12] 侦测次数
            ],
            alloc=alloc, ctx=ctx, desc="子弹通用-碰撞[主角]（DELAY 后挂载）",
        )
        extra_nodes.append(collision_self_node)

        # 1g) DELAY 节点：飞出+悬停 +5 帧后才挂第二碰撞模板（避免出生即死）
        delay_collision_self_node = Node(
            guid=alloc.new_guid(),
            cls="TSET_DELAY_EXECUTE",
            config_id=alloc.allocate_effect_id(),
            desc=f"DELAY {flight_frames + hover_frames + 5} 帧后挂主角检测",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_DELAY_EXECUTE"),
                "Params": [
                    TParam(value=flight_frames + hover_frames + 5).to_dict(),  # 帧数
                    TParam(value=collision_self_node.config_id).to_dict(),     # 内嵌节点
                    TParam(value=0).to_dict(),  # die_continue
                    TParam(value=0).to_dict(),  # 筛选 ID
                    TParam(value=0).to_dict(),  # 中断 ID
                    TParam(value=0).to_dict(),  # 保留
                    TParam(value=1).to_dict(),  # rapid_affect
                ],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_DELAY_EXECUTE")},
            table_name="SkillEffectConfig",
        )
        delay_collision_self_node.config_payload["ID"] = delay_collision_self_node.config_id
        extra_nodes.append(delay_collision_self_node)

        # 1h) AfterBorn ORDER：[第一碰撞模板（敌方）+ DELAY 包的第二碰撞模板（主角）]
        after_born_order_node = Node(
            guid=alloc.new_guid(),
            cls="TSET_ORDER_EXECUTE",
            config_id=alloc.allocate_effect_id(),
            desc="AfterBorn ORDER：敌方碰撞 + DELAY 后挂主角碰撞",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
                "Params": [
                    TParam(value=collision_enemy_node.config_id).to_dict(),
                    TParam(value=delay_collision_self_node.config_id).to_dict(),
                ],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
            table_name="SkillEffectConfig",
        )
        after_born_order_node.config_payload["ID"] = after_born_order_node.config_id
        extra_nodes.append(after_born_order_node)
        bullet_after_born_id = after_born_order_node.config_id

    # === 2) Bullet 配置（v1.2.4 修正 — 取消距离截断，让回飞充足时间）===
    # PostMortem #010：v1.2.3 LifeFlag=3 + MaxDistance=max_distance+200 导致
    #   累计飞行（飞出 + 回飞一半）就触发距离死亡，子弹"还没碰到主角就消失"。
    # v1.2.4 修正：
    #   ① LifeFlag = 1（仅时间触发死亡，不让距离截断）
    #   ② MaxDistance = max_distance * 4（巨大余量，避免任何距离截断）
    #   ③ LastTime 算够：flight + hover + 回飞时间（按 60 fps 换算）+ 10 帧缓冲
    #      回飞时间（帧）= max_distance / return_speed × 60
    return_frames_estimate = max(15, int(max_distance * 60 / max(return_speed, 1)))
    bullet_total_lifetime = flight_frames + hover_frames + return_frames_estimate + 10
    bullet_cfg_node = _make_inline_bullet_config_node(
        bullet_id=bullet1_id,
        fly_type=5,
        trace_path_type=3,
        trace_path_params=[25, 8, max_distance, 0, return_speed, 0],
        speed=2000,
        last_time=bullet_total_lifetime,
        max_distance=max_distance * 4,                                   # ② 巨大余量
        after_born_id=bullet_after_born_id,
        die_skill_id=0,
        alloc=alloc, ctx=ctx,
        desc=f"回旋飞回 Bullet（远距={max_distance}, 回飞估{return_frames_estimate}帧, 总寿命={bullet_total_lifetime}）",
    )
    bullet_cfg_node.config_payload["AcceSpeed"] = 1500
    bullet_cfg_node.config_payload["MaxSpeed"] = max(5000, return_speed * 2)
    bullet_cfg_node.config_payload["ChaseTargetEnemy_FaceToTarget"] = True
    bullet_cfg_node.config_payload["ChaseTargetEnemy_PitchFaceToTarget"] = True
    bullet_cfg_node.config_payload["LifeFlag"] = 1                       # ① 仅时间触发死亡
    extra_nodes.append(bullet_cfg_node)

    # === 5) N 颗 Bullet1 创建（按 angles 角度分布）===
    # 复用 v1.1 angles 多颗逻辑：caster.facing 节点 + N-1 个 facing+offset CALC 节点
    facing_node: Node | None = None
    if angles and any(a != 0 for a in angles):
        facing_node = _make_get_caster_facing_node(alloc)
        extra_nodes.append(facing_node)

    bullet_create_nodes: list[Node] = []
    facing_calc_nodes: list[Node] = []
    if not angles:
        angles = [0]   # 单颗时也走相同路径

    for offset in angles:
        if offset == 0:
            facing_param = resolve_ref("attr:面向")
        else:
            calc_node = _make_facing_offset_calc_node(facing_node.config_id, offset, alloc)
            facing_calc_nodes.append(calc_node)
            facing_param = TParam(value=calc_node.config_id, param_type=2)

        bn = Node(
            guid=alloc.new_guid(),
            cls="TSET_CREATE_BULLET",
            config_id=alloc.allocate_effect_id(),
            desc=f"创建 Bullet1 飞出 (offset={offset}°)" if offset else "创建 Bullet1 飞出",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_BULLET"),
                "Params": [
                    TParam(value=bullet1_id).to_dict(),
                    facing_param.to_dict(),
                    resolve_ref("attr:位置X").to_dict(),
                    resolve_ref("attr:位置Y").to_dict(),
                    creator.to_dict(),
                    TParam(value=offset_right).to_dict(),
                    TParam(value=offset_forward).to_dict(),
                    TParam(value=0).to_dict(), TParam(value=0).to_dict(),
                    TParam(value=1).to_dict(),  # is_projectile
                    TParam(value=0).to_dict(), TParam(value=0).to_dict(),
                    TParam(value=z_height).to_dict(),
                    TParam(value=0).to_dict(),
                    TParam(value=rapid_affect).to_dict(),
                ],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_BULLET")},
            table_name="SkillEffectConfig",
        )
        bn.config_payload["ID"] = bn.config_id
        bullet_create_nodes.append(bn)
    extra_nodes.extend(facing_calc_nodes)

    # === 6) 顶层 ORDER 把 N 颗 Bullet1 串起来 ===
    fan_order_node = Node(
        guid=alloc.new_guid(),
        cls="TSET_ORDER_EXECUTE",
        config_id=alloc.allocate_effect_id(),
        desc=f"回旋飞回 {len(bullet_create_nodes)} 颗（angles={angles}）",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
            "Params": [TParam(value=n.config_id).to_dict() for n in bullet_create_nodes],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
        table_name="SkillEffectConfig",
    )
    fan_order_node.config_payload["ID"] = fan_order_node.config_id

    # 返回顺序：fan_order 作为代表节点（被父级 delay/order 引用）
    return [fan_order_node] + bullet_create_nodes + extra_nodes


# ------------------------------------------------------------
# 通用 step expander
# ------------------------------------------------------------
def expand_cast_anim(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    cfg = step["cast_anim"]
    entity = resolve_ref(cfg.get("entity", "entity:主体"))
    params = [
        entity,
        TParam(value=cfg["anim_id"]),
        TParam(value=1 if cfg.get("upper_body", False) else 0),
        TParam(value=cfg.get("blend_ms", 200)),
        TParam(value=cfg.get("speed_pct", 100)),
        TParam(value=1 if cfg.get("movable_break", True) else 0),
        TParam(value=1 if cfg.get("rapid_affect", False) else 0),
        TParam(value=0),  # 仅指定可见
        TParam(value=0),  # 帧数
        TParam(value=0),  # 音效绑定
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_PLAY_ROLE_ANIM",
        config_id=alloc.allocate_effect_id(),
        desc=cfg.get("desc", "播放动作"),
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_PLAY_ROLE_ANIM"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_PLAY_ROLE_ANIM")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return [n]


def expand_cast_effect(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    cfg = step["cast_effect"]
    follow = resolve_ref(cfg.get("follow", "entity:主体"))
    params = [
        TParam(value=cfg["model_id"]),
        resolve_ref("attr:面向"),
        resolve_ref("attr:位置X"),
        resolve_ref("attr:位置Y"),
        TParam(value=cfg.get("duration_frames", 30)),
        follow,
        TParam(value=cfg.get("scale_pct", 100)),
        TParam(value=cfg.get("destroy_delay_ms", 1000)),
        TParam(value=0),  # 偏移X
        TParam(value=cfg.get("speed_pct", 100)),
        TParam(value=0),  # 偏移Y
        TParam(value=0),  # 单位组
        TParam(value=cfg.get("z_height", 0)),
        TParam(value=0),  # 特效类型
        TParam(value=0),  # 占位
        TParam(value=0),  # 出生后效果
        TParam(value=1 if cfg.get("rapid_affect", True) else 0),
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_CREATE_EFFECT",
        config_id=alloc.allocate_effect_id(),
        desc=cfg.get("desc", "播放特效"),
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_EFFECT"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_EFFECT")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id

    # 同时生成 ModelConfigNode 引用节点（让策划在编辑器中能直观看到模型路径）
    model_cfg_node = make_model_config_node(cfg["model_id"], alloc, ctx,
                                             desc=cfg.get("desc", ""))
    return [n, model_cfg_node]


def expand_delay(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    cfg = step["delay"]
    nested_nodes: list[Node] = []
    nested_root_id = 0

    inner_steps = cfg["then"]
    if len(inner_steps) == 1:
        # 直接指向单个内部节点
        inner = expand_step(inner_steps[0], alloc, ctx)
        nested_nodes.extend(inner)
        nested_root_id = inner[0].config_id
    else:
        # 多个 → 用 ORDER_EXECUTE 包起来
        order_node, children = build_order(inner_steps, alloc, ctx, desc="(delay group)")
        nested_nodes.append(order_node)
        nested_nodes.extend(children)
        nested_root_id = order_node.config_id

    frames_param = resolve_ref(cfg["frames"])
    params = [
        frames_param,
        TParam(value=nested_root_id),
        TParam(value=1 if cfg.get("die_continue", False) else 0),
        TParam(value=0),  # 筛选 ID
        TParam(value=0),  # 中断 ID
        TParam(value=0),  # 保留
        TParam(value=1 if cfg.get("rapid_affect", True) else 0),
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_DELAY_EXECUTE",
        config_id=alloc.allocate_effect_id(),
        desc=cfg.get("desc", "延迟执行"),
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_DELAY_EXECUTE"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_DELAY_EXECUTE")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return [n] + nested_nodes


def expand_play_sound(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    cfg = step["play_sound"]
    params = [
        resolve_ref(cfg.get("voice_id", "entity:主体伤害归属")),
        TParam(value=cfg["sound_type"]),
        TParam(value=cfg.get("config", 1)),
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_PLAY_SOUND",
        config_id=alloc.allocate_effect_id(),
        desc="播放音效",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_PLAY_SOUND"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_PLAY_SOUND")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return [n]


def expand_camera_shake(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    cfg = step["camera_shake"]
    params = [
        resolve_ref(cfg.get("executor", "entity:施法者根")),
        TParam(value=int(cfg.get("target", 1))),
        TParam(value=cfg.get("radius", 800)),
        TParam(value=0),  # 抖X
        TParam(value=0),  # 抖Y
        TParam(value=cfg["shake_id"]),
        TParam(value=cfg.get("priority", 1)),
        TParam(value=0),  # 插入方式
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_CAMERA_SHAKE",
        config_id=alloc.allocate_effect_id(),
        desc="镜头抖动",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CAMERA_SHAKE"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CAMERA_SHAKE")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return [n]


def expand_apply_buff(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    cfg = step["apply_buff"]
    params = [
        resolve_ref(cfg.get("host", "entity:目标")),
        resolve_ref(cfg.get("source", "entity:主体")),
        TParam(value=cfg["buff_id"]),
        TParam(value=cfg.get("duration_frames", 0)),
        TParam(value=cfg.get("interval_frames", 0)),
        TParam(value=cfg.get("layers", 1)),
        TParam(value=1 if cfg.get("rapid_affect", True) else 0),
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_ADD_BUFF",
        config_id=alloc.allocate_effect_id(),
        desc="添加 Buff",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ADD_BUFF"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ADD_BUFF")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return [n]


def expand_remove_buff(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    cfg = step["remove_buff"]
    params = [
        resolve_ref(cfg.get("target", "entity:目标")),
        TParam(value=0),  # 移除方式：BUFF_ID
        TParam(value=cfg["buff_id"]),
        TParam(value=0),
        TParam(value=0),
        TParam(value=0),
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_REMOVE_BUFF",
        config_id=alloc.allocate_effect_id(),
        desc="移除 Buff",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_REMOVE_BUFF"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_REMOVE_BUFF")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return [n]


def expand_modify_tag(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    cfg = step["modify_tag"]
    params = [
        resolve_ref(cfg.get("owner", "entity:主体")),
        resolve_ref(cfg.get("skill_id", "0")),
        TParam(value=cfg["tag_id"]),
        TParam(value=cfg["value"]),
        TParam(value=1),  # 参数类型
    ]
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_MODIFY_SKILL_TAG_VALUE",
        config_id=alloc.allocate_effect_id(),
        desc="修改技能参数",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_MODIFY_SKILL_TAG_VALUE"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_MODIFY_SKILL_TAG_VALUE")},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return [n]


# ============================================================
# v2.0 新增：M2-M7 expander
# ============================================================
def expand_displacement(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """v2.0 M2: 高速位移 → 调用 175_0023 位移_按速度距离 模板。

    模板 13 个 TemplateParam（已校对 GATE-0.5 §3）：
      [0] 位移速度
      [1] 位移距离
      [2] 方向类型 (TSkillDirectionType: 0=固定方向 1=移动方向 2=面向)
      [3] 偏移角度（用于固定方向）
      [4] 是否穿越中间阻挡
      [5] 是否落点挡在阻挡外
      [6] 受到控制是否停止位移
      [7] 是否穿越角色
      [8] 位移中执行效果（每帧）— SkillEffectConfig
      [9] 位移后执行效果
      [10] 位移中断执行效果
      [11] 位移中断自定义条件 — SkillConditionConfig
      [12] 是否忽略定身状态
    """
    cfg = step["displacement"]
    direction = cfg.get("direction", "朝向指示器")
    direction_map = {"朝向指示器": 2, "面向": 2, "朝向目标": 2}  # 简化：都用面向
    direction_int = direction_map.get(direction, 2) if isinstance(direction, str) else 0
    offset_angle = direction if isinstance(direction, int) else 0

    return [make_template_call_node(
        template_key="位移_按速度距离",
        extra_params=[
            cfg["speed"],                    # [0] 位移速度
            cfg["distance"],                 # [1] 位移距离
            direction_int,                   # [2] 方向类型 = 2 面向
            offset_angle,                    # [3] 偏移角度
            1,                               # [4] 穿越中间阻挡
            1,                               # [5] 落点挡在阻挡外
            0 if cfg.get("invulnerable", True) else 1,  # [6] 控制是否停止
            1,                               # [7] 穿越角色
            0,                               # [8] 位移中执行（暂不支持 hit_along_path）
            0,                               # [9] 位移后执行
            0,                               # [10] 中断执行
            0,                               # [11] 中断条件
            1 if cfg.get("invulnerable", True) else 0,  # [12] 忽略定身
        ],
        alloc=alloc, ctx=ctx,
        desc=cfg.get("desc", "高速位移"),
    )]


def expand_summon_clone(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """v2.0 M3 + M4: 召唤残影 → 调用 146004117 召唤魂影单位 模板。

    模板 9 个 TemplateParam（GATE-0.5 §3 已校对）：
      [0] 面向
      [1] 位置X
      [2] 位置Y
      [3] 延迟帧数自动死亡（0=不自动死亡）
      [4] 单位组 — TSkillEntityGroupType (4=召唤单位1)
      [5] 出生前额外执行效果 — SkillEffectConfig
      [6] 出生后额外执行效果 — SkillEffectConfig（残影主动行为挂这里）
      [7] 魂影释放技能前是否重新继承召唤者属性
      [8] 魂影继承属性比例（万分比）
    """
    cfg = step["summon_clone"]
    duration = cfg["duration_frames"]
    extra_nodes: list[Node] = []

    # M4 PoC: active_pattern → 出生后执行的"周期搜敌发弹" ORDER
    # 实现机制（PoC）：用 TSET_REPEAT_EXECUTE 周期触发 → 内嵌 bullet step
    after_born_id = 0
    active = cfg.get("active_pattern")
    if active:
        # 把 on_target_found 的 step 列表展开
        target_steps = active.get("on_target_found") or []
        sub_nodes: list[Node] = []
        sub_root_id = 0
        if target_steps:
            if len(target_steps) == 1:
                sub = expand_step(target_steps[0], alloc, ctx)
                sub_nodes.extend(sub)
                sub_root_id = sub[0].config_id
            else:
                order_node, children = build_order(target_steps, alloc, ctx, desc="残影主动行为子流程")
                sub_nodes.append(order_node)
                sub_nodes.extend(children)
                sub_root_id = order_node.config_id
        extra_nodes.extend(sub_nodes)

        # REPEAT 节点（周期触发）
        repeat_eid = alloc.allocate_effect_id()
        repeat_payload = {
            "ID": repeat_eid,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_REPEAT_EXECUTE") or 3,
            "Params": [
                TParam(value=active.get("interval_frames", 60)).to_dict(),  # [0] 间隔帧
                TParam(value=duration // active.get("interval_frames", 60) + 1).to_dict(),  # [1] 次数
                TParam(value=0).to_dict(),                                    # [2] 立即执行（0=否，等间隔）
                TParam(value=sub_root_id).to_dict(),                          # [3] 子效果
                TParam(value=0).to_dict(), TParam(value=0).to_dict(),
                TParam(value=0).to_dict(), TParam(value=0).to_dict(),
                TParam(value=0).to_dict(),
                TParam(value=1).to_dict(),                                    # 急速
            ],
        }
        repeat_node = Node(
            guid=alloc.new_guid(),
            cls="TSET_REPEAT_EXECUTE",
            config_id=repeat_eid,
            desc=f"残影周期搜敌发弹（每{active.get('interval_frames',60)}帧）",
            config_payload=repeat_payload,
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_REPEAT_EXECUTE") or 3},
            table_name="SkillEffectConfig",
        )
        extra_nodes.append(repeat_node)
        after_born_id = repeat_eid

    # 召唤魂影 模板调用
    summon_node = make_template_call_node(
        template_key="召唤魂影单位",
        extra_params=[
            "attr:面向",                    # [0] 面向 = caster.facing
            "attr:位置X",                   # [1] 位置X
            "attr:位置Y",                   # [2] 位置Y
            duration,                       # [3] 延迟帧数自动死亡
            4,                              # [4] 单位组 = 召唤单位1
            0,                              # [5] 出生前
            after_born_id,                  # [6] 出生后（active_pattern 的 REPEAT）
            1,                              # [7] 重新继承属性
            10000,                          # [8] 继承比例 100%
        ],
        alloc=alloc, ctx=ctx,
        desc=cfg.get("desc", f"召唤残影（{duration}帧）"),
    )

    # 记录残影位置到 SkillTag（v2.2 真实现）
    # record_position_to_tag=N → 记录 caster X 到 tag N，记录 caster Y 到 tag N+1
    # second_stage 的 return_to_clone 用 tag:N / tag:N+1 取回（PT=3 SKILL_PARAM）
    record_tag = cfg.get("record_position_to_tag", 0)
    if record_tag:
        record_nodes = _make_record_caster_position_nodes(record_tag, alloc, ctx)
        record_order = record_nodes[0]  # ORDER 是首项
        ctx.info(f"M3 v2.2: 记录 caster 位置 X→tag {record_tag} / Y→tag {record_tag+1}（{len(record_nodes)} 节点）")

        # 用一个外层 ORDER 同时调度 [position-record, summon] —— 都需要执行
        wrap_order = Node(
            guid=alloc.new_guid(),
            cls="TSET_ORDER_EXECUTE",
            config_id=alloc.allocate_effect_id(),
            desc="位置记录 + 召唤残影",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
                "Params": [
                    TParam(value=record_order.config_id).to_dict(),
                    TParam(value=summon_node.config_id).to_dict(),
                ],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
            table_name="SkillEffectConfig",
        )
        wrap_order.config_payload["ID"] = wrap_order.config_id
        # wrap_order 作为代表节点，子节点全部跟随
        return [wrap_order, summon_node] + extra_nodes + record_nodes

    return [summon_node] + extra_nodes


def _make_record_caster_position_nodes(
    base_tag_id: int, alloc: IdAllocator, ctx: "BuildCtx",
) -> list[Node]:
    """v2.2：生成 4 个辅助节点 — 记录 caster X/Y 到 SkillTags。

    返回节点列表（首项不是代表节点，全部都需被父级 ORDER 引用，所以
    expand_summon_clone 会把这几个节点 extra_nodes 处理为同级邻居）。

    生成结构：
      GET_ATTR(caster, 位置X=59) ── value ──→ MODIFY_TAG(skill_id_inst=41, tag=base, value=PT=2 effect_return)
      GET_ATTR(caster, 位置Y=60) ── value ──→ MODIFY_TAG(skill_id_inst=41, tag=base+1, value=PT=2 effect_return)

    实际执行需要这 4 个节点都被 ORDER 顺序调用。但当前 expand_summon_clone
    把它们丢到 extra_nodes 里，没有 ORDER 包裹——这里直接返回 ORDER + 4 子节点。
    """
    # GET caster.位置X (TBattleNatureEnum 59)
    get_x = Node(
        guid=alloc.new_guid(),
        cls="TSET_GET_ENTITY_ATTR_VALUE",
        config_id=alloc.allocate_effect_id(),
        desc="读 caster 位置X",
        config_payload={
            "ID": 0,
            "SkillEffectType": 32,
            "Params": [
                resolve_ref("entity:施法者").to_dict(),  # [0] 单位
                TParam(value=59).to_dict(),               # [1] 属性 = 位置X
            ],
        },
        extra_data={"SkillEffectType": 32},
        table_name="SkillEffectConfig",
    )
    get_x.config_payload["ID"] = get_x.config_id

    get_y = Node(
        guid=alloc.new_guid(),
        cls="TSET_GET_ENTITY_ATTR_VALUE",
        config_id=alloc.allocate_effect_id(),
        desc="读 caster 位置Y",
        config_payload={
            "ID": 0,
            "SkillEffectType": 32,
            "Params": [
                resolve_ref("entity:施法者").to_dict(),
                TParam(value=60).to_dict(),               # 位置Y
            ],
        },
        extra_data={"SkillEffectType": 32},
        table_name="SkillEffectConfig",
    )
    get_y.config_payload["ID"] = get_y.config_id

    # MODIFY tag base = X
    # Params: [entity=3 caster PT=5, skill="-"=0/PT=0 (实体级 tag,跨技能), tag_id PT=0, value PT=2, is_modify=1]
    # v2.3.2：Param[1] 改为 PT=0/Value=0（"-" 类型）—— tag 绑定在实体上，不绑技能；
    # 第二段技能直接读同 tag 即可，无需 source_skill_id
    set_x = Node(
        guid=alloc.new_guid(),
        cls="TSET_MODIFY_SKILL_TAG_VALUE",
        config_id=alloc.allocate_effect_id(),
        desc=f"存 X → 实体 tag {base_tag_id}",
        config_payload={
            "ID": 0,
            "SkillEffectType": 46,
            "Params": [
                TParam(value=3, param_type=5).to_dict(),               # [0] 实体 = 施法者
                TParam(value=0, param_type=0).to_dict(),               # [1] 技能 = "-"（实体级 tag）
                TParam(value=base_tag_id).to_dict(),                   # [2] tag id
                TParam(value=get_x.config_id, param_type=2).to_dict(), # [3] value = effect_return
                TParam(value=1).to_dict(),                              # [4] 设置（不是叠加）
            ],
        },
        extra_data={"SkillEffectType": 46},
        table_name="SkillEffectConfig",
    )
    set_x.config_payload["ID"] = set_x.config_id

    set_y = Node(
        guid=alloc.new_guid(),
        cls="TSET_MODIFY_SKILL_TAG_VALUE",
        config_id=alloc.allocate_effect_id(),
        desc=f"存 Y → 实体 tag {base_tag_id+1}",
        config_payload={
            "ID": 0,
            "SkillEffectType": 46,
            "Params": [
                TParam(value=3, param_type=5).to_dict(),
                TParam(value=0, param_type=0).to_dict(),
                TParam(value=base_tag_id + 1).to_dict(),
                TParam(value=get_y.config_id, param_type=2).to_dict(),
                TParam(value=1).to_dict(),
            ],
        },
        extra_data={"SkillEffectType": 46},
        table_name="SkillEffectConfig",
    )
    set_y.config_payload["ID"] = set_y.config_id

    # ORDER 包裹 set_x + set_y（get_x/get_y 通过 effect_return 被 set_x/set_y 引用，自动随之执行）
    order_node = Node(
        guid=alloc.new_guid(),
        cls="TSET_ORDER_EXECUTE",
        config_id=alloc.allocate_effect_id(),
        desc=f"记录 caster 位置 → tag {base_tag_id}/{base_tag_id+1}",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
            "Params": [
                TParam(value=set_x.config_id).to_dict(),
                TParam(value=set_y.config_id).to_dict(),
            ],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
        table_name="SkillEffectConfig",
    )
    order_node.config_payload["ID"] = order_node.config_id

    return [order_node, set_x, set_y, get_x, get_y]


# sub_type 中文名 → TSkillSlotType 整数值（默认槽位推断表）
_SUB_TYPE_TO_SLOT = {
    "功法技": 1,        # TSST_GONG_FA
    "功法": 1,
    "奇术":  2,         # TSST_QISHU_1（默认奇术 1）
    "奇术1": 2,
    "奇术2": 3,         # TSST_QISHU_2
    "神通":  4,         # TSST_SHEN_TONG
}

# swap_slot 中文别名 → 整数值（IR 用户层）
_SWAP_SLOT_CN_TO_INT = {
    "功法":  1,
    "奇术1": 2,
    "奇术2": 3,
    "神通":  4,
    "额外1": 5,
}


def _resolve_swap_slot(cfg: dict, ir: dict) -> int | str:
    """解析 swap_slot：'动态'（默认）→ 返回 sentinel 字符串 'DYNAMIC'，
    显式值（中文/整数）→ 返回 int。"""
    explicit = cfg.get("swap_slot", "动态")
    if isinstance(explicit, int):
        return explicit
    if explicit == "动态":
        return "DYNAMIC"
    if isinstance(explicit, str):
        v = _SWAP_SLOT_CN_TO_INT.get(explicit)
        if v is not None:
            return v
        # fallback to sub_type 推断
        v = _SUB_TYPE_TO_SLOT.get(explicit)
        if v is not None:
            return v
        raise CompileError(f"未识别的 swap_slot: {explicit}")
    raise CompileError(f"swap_slot 类型不支持: {type(explicit).__name__}")


def _make_get_skill_slot_node(alloc: IdAllocator, ctx: "BuildCtx") -> Node:
    """v2.2：生成 TSET_GET_SKILL_SLOT_TYPE 辅助节点 — 取施法者当前装备的初始技能槽位。

    Params（与 30221000 rid=1080 一致）：
      [0] 技能实例 = 41/PT=5 (TCommonParamType.TCPT_ORIGIN_SKILL_INST_ID 初始技能实例ID)
      [1] 0/PT=0 — 占位
    """
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_GET_SKILL_SLOT_TYPE",
        config_id=alloc.allocate_effect_id(),
        desc="取施法者当前技能槽位",
        config_payload={
            "ID": 0,
            "SkillEffectType": 260,
            "Params": [
                TParam(value=41, param_type=5).to_dict(),
                TParam(value=0).to_dict(),
            ],
        },
        extra_data={"SkillEffectType": 260},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


def expand_two_stage_skill(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """v2.1 两段式技能（按键真机制）：

    设计依据：30221000 反向工程发现"两段式按键"在项目中是用 44014633 修改单位槽位模板
    实现的——把当前槽位的技能 ID 换成第二段独立 SkillConfig ID（如 30221002），玩家再按
    同一按钮时实际触发的就是第二段。175_0102 技能连招(2段) 是"自动连段"，不是按键判断。

    实现方式：
      - first_stage 的 step 内联到 ORDER
      - 末尾追加 44014633 调用：把当前槽位换成 second_stage_skill_id
      - second_stage 由 compile_ir 顶层流程独立编译为 second JSON

    44014633 模板 7 个 TemplateParam：
      [0] 单位ID — 施法者
      [1] 技能槽位 — TSkillSlotType（按 sub_type 推断或 swap_slot 显式指定）
      [2] 技能ID — second_stage_skill_id（独立 SkillConfig）
      [3] 技能等级 — 1（PoC，等 v2.2 接当前等级查询）
      [4] 持续时间 — auto_reset_frames（超时自动还原）
      [5] 切回返还百分比CD — swap_cd_refund_pct
      [6] 显示修改后技能持续时间UI — swap_show_ui
    """
    cfg = step["two_stage_skill"]
    extra_nodes: list[Node] = []

    # 1. 内联编译 first_stage 的 step
    first_root_ids: list[int] = []
    for fs_step in cfg["first_stage"]:
        sub = expand_step(fs_step, alloc, ctx)
        first_root_ids.append(sub[0].config_id)
        extra_nodes.extend(sub)

    # 2. 解析 swap 参数
    main_skill_id = ctx.skill_id
    second_skill_id = cfg.get("second_stage_skill_id", main_skill_id + 1)
    swap_slot = _resolve_swap_slot(cfg, ctx.ir)
    auto_reset = cfg.get("auto_reset_frames", 600)
    swap_refund = cfg.get("swap_cd_refund_pct", 100)
    swap_show_ui = 1 if cfg.get("swap_show_ui", True) else 0

    # v2.2：默认 swap_slot='动态'，生成 TSET_GET_SKILL_SLOT_TYPE 辅助节点
    if swap_slot == "DYNAMIC":
        slot_node = _make_get_skill_slot_node(alloc, ctx)
        extra_nodes.append(slot_node)
        slot_param = TParam(value=slot_node.config_id, param_type=2)  # PT=2 effect_return
        ctx.info(f"v2.2: swap_slot=动态 → 生成 GET_SKILL_SLOT_TYPE 节点 ID={slot_node.config_id}")
    else:
        slot_param = swap_slot  # 整数常量 → resolve_ref 处理为 PT=0

    # 3. 末尾插入 44014633 调用
    swap_node = make_template_call_node(
        template_key="修改单位槽位",
        extra_params=[
            "entity:施法者",     # [3] 单位ID = TCPT_CREATE_ENTITY (3)
            slot_param,          # [4] 技能槽位（v2.2 默认 PT=2 effect_return 取动态槽位）
            second_skill_id,     # [5] 技能ID（PT=0 常量；指向独立 SkillConfig）
            1,                   # [6] 技能等级（PoC：固定 1）
            auto_reset,          # [7] 持续时间
            swap_refund,         # [8] 切回返还百分比CD
            swap_show_ui,        # [9] 显示UI
        ],
        alloc=alloc, ctx=ctx,
        desc=f"切槽位至第二段 SkillConfig {second_skill_id}",
    )
    first_root_ids.append(swap_node.config_id)
    extra_nodes.append(swap_node)

    # 4. ORDER 包裹 first_stage 各 step + swap 调用
    params = [TParam(value=cid) for cid in first_root_ids]
    order_node = Node(
        guid=alloc.new_guid(),
        cls="TSET_ORDER_EXECUTE",
        config_id=alloc.allocate_effect_id(),
        desc="第一段（首次按键）+ 切槽位",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
        table_name="SkillEffectConfig",
    )
    order_node.config_payload["ID"] = order_node.config_id
    return [order_node] + extra_nodes


def expand_return_to_clone(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """v2.0 M6: 飞回残影位置 → 调用 175_0024 位移_按速度目标点 模板。

    模板 11 个 TemplateParam：
      [0] 位移速度
      [1] 目标点_X
      [2] 目标点_Y
      [3..6] 阻挡 / 控制 / 角色 / 位移中执行
      [7] 位移后执行（落地 AOE 挂这里）
      [8..10] 中断 / 忽略定身
    """
    cfg = step["return_to_clone"]
    extra_nodes: list[Node] = []

    # 落地特效（可选）
    landing_effect_id = 0
    le = cfg.get("landing_effect")
    if le and le.get("model_id"):
        eff_step = {"cast_effect": {
            "model_id": le["model_id"],
            "scale_pct": le.get("scale_pct", 100),
            "duration_frames": 30,
            "desc": "落地特效",
        }}
        eff_nodes = expand_cast_effect(eff_step, alloc, ctx)
        landing_effect_id = eff_nodes[0].config_id
        extra_nodes.extend(eff_nodes)

    # v2.3.2：实体级 SkillTag —— 不绑技能，跨技能透明读取
    # tag 由第一段 MODIFY_SKILL_TAG_VALUE([entity=3, skill="-"=0/PT=0, ...]) 写入
    # 这里 GET_SKILL_TAG_VALUE 用同样的 [skill=0/PT=0] 即可读到
    base_tag = cfg["target_position_tag"]
    get_tag_x = _make_get_skill_tag_value_node(base_tag,     alloc, desc=f"读实体 tag {base_tag} (X)")
    get_tag_y = _make_get_skill_tag_value_node(base_tag + 1, alloc, desc=f"读实体 tag {base_tag + 1} (Y)")
    extra_nodes.extend([get_tag_x, get_tag_y])

    move_node = make_template_call_node(
        template_key="位移_按速度目标点",
        extra_params=[
            cfg["speed"],                                                      # [0] 位移速度
            TParam(value=get_tag_x.config_id, param_type=2),                  # [1] X = effect_return GET_SKILL_TAG_VALUE
            TParam(value=get_tag_y.config_id, param_type=2),                  # [2] Y = effect_return GET_SKILL_TAG_VALUE
            1,                              # [3] 穿越阻挡
            1,                              # [4] 落点挡在阻挡外
            0 if cfg.get("invulnerable", True) else 1,  # [5] 控制停止
            0,                              # [6] 位移中执行
            landing_effect_id,              # [7] 位移后执行（落地特效）
            0,                              # [8] 中断执行
            0,                              # [9] 中断条件
            1 if cfg.get("invulnerable", True) else 0,  # [10] 忽略定身
        ],
        alloc=alloc, ctx=ctx,
        desc=cfg.get("desc", f"飞回残影位置（读实体 tag {base_tag}/{base_tag+1}）"),
    )
    return [move_node] + extra_nodes


def _make_get_skill_tag_value_node(
    tag_id: int,
    alloc: IdAllocator,
    desc: str = "",
) -> Node:
    """v2.3.2：生成 TSET_GET_SKILL_TAG_VALUE 节点（实体级 tag，跨技能透明读取）。

    Params（与 30221000/30221002 Pattern B 一致）：
      [0] V=3/PT=5  — 施法者实例ID（tag 挂在 caster 身上）
      [1] V=0/PT=0  — 技能 "-"（实体级 tag，不绑特定技能；写入和读取都用此）
      [2] V=<tag_id>/PT=0 — tag ID
      [3] V=1/PT=0  — 取最终值（含 modifiers）
      [4] V=0/PT=0  — 占位
    """
    n = Node(
        guid=alloc.new_guid(),
        cls="TSET_GET_SKILL_TAG_VALUE",
        config_id=alloc.allocate_effect_id(),
        desc=desc or f"读实体 tag {tag_id}",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_GET_SKILL_TAG_VALUE") or 48,
            "Params": [
                TParam(value=3, param_type=5).to_dict(),
                TParam(value=0, param_type=0).to_dict(),
                TParam(value=tag_id).to_dict(),
                TParam(value=1).to_dict(),
                TParam(value=0).to_dict(),
            ],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_GET_SKILL_TAG_VALUE") or 48},
        table_name="SkillEffectConfig",
    )
    n.config_payload["ID"] = n.config_id
    return n


def expand_aoe_circle(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """v2.0 M7: 圆形 AOE → 用 子弹通用-碰撞模板（圆形 + 自身位置）的轻量化变体。

    实现选择：
      复用子弹通用-碰撞模板（target_camp + 半径）+ 子弹通用-伤害模板，
      但不挂在子弹身上，而是 caster 位置触发一次。
    简化做法：直接调用伤害模板对 entity:目标 造伤害（伤害模板内部会按筛选目标循环）。

    PoC 阶段 AOE 简化为：在主体位置造一次直接伤害（半径不参与筛选）。
    完整 AOE 多目标筛选留 v2.1。
    """
    cfg = step["aoe_circle"]
    damage_cfg = cfg.get("damage") or {}
    elem_map = {"随技能": 0, "金": 1, "木": 2, "水": 3, "火": 4, "土": 5}
    type_map = {"随技能": 0, "物理": 1, "法术": 2, "真实": 3}
    d_elem = damage_cfg.get("element", "随技能")
    d_elem_int = elem_map.get(d_elem, 0) if isinstance(d_elem, str) else int(d_elem)
    d_type = damage_cfg.get("type", "随技能")
    d_type_int = type_map.get(d_type, 0) if isinstance(d_type, str) else int(d_type)

    extra_nodes: list[Node] = []

    # 落地特效（可选）
    landing_effect_id = 0
    le = cfg.get("landing_effect")
    if le and le.get("model_id"):
        eff_step = {"cast_effect": {
            "model_id": le["model_id"],
            "scale_pct": le.get("scale_pct", 100),
            "duration_frames": 30,
            "desc": "AOE 落地特效",
        }}
        eff_nodes = expand_cast_effect(eff_step, alloc, ctx)
        landing_effect_id = eff_nodes[0].config_id
        extra_nodes.extend(eff_nodes)

    # 调用子弹通用-伤害模板（PoC: 单目标伤害）
    damage_node = make_template_call_node(
        template_key="子弹通用逻辑-伤害",
        extra_params=[
            1, "entity:施法者伤害归属",
            d_type_int, d_elem_int,
            damage_cfg.get("subtype_flags", 16),
            damage_cfg.get("coef", 12000),
            damage_cfg.get("extra", -1),
            0, 0, 0,
            0,                              # 命中后不销毁
            -1, -1,
        ],
        alloc=alloc, ctx=ctx,
        desc=f"AOE 圆形伤害 R={cfg['radius']} (PoC 简化版)",
    )
    extra_nodes.append(damage_node)

    # 用 ORDER 包代表节点（修 PostMortem #014：避免 damage_node 孤立）
    children_ids = [damage_node.config_id]
    if landing_effect_id:
        children_ids = [landing_effect_id] + children_ids
    aoe_order = Node(
        guid=alloc.new_guid(),
        cls="TSET_ORDER_EXECUTE",
        config_id=alloc.allocate_effect_id(),
        desc=f"AOE 圆形 R={cfg['radius']} ORDER",
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
            "Params": [TParam(value=cid).to_dict() for cid in children_ids],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
        table_name="SkillEffectConfig",
    )
    aoe_order.config_payload["ID"] = aoe_order.config_id

    return [aoe_order] + extra_nodes


def expand_aoe_bullet_circle(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """圆形AOE落点子弹：FlyType=0静态子弹挂碰撞模板，命中范围内所有敌方目标。"""
    cfg = step["bullet"]
    bullet_id = cfg["bullet_id"]
    hit_cfg = cfg.get("hit") or {}
    offset_forward = cfg.get("offset_forward", 0)
    offset_right = cfg.get("offset_right", 0)
    extra_nodes: list[Node] = []
    after_born_id = 0

    if hit_cfg:
        elem_map = {"随技能": 0, "金": 1, "木": 2, "水": 3, "火": 4, "土": 5}
        type_map = {"随技能": 0, "物理": 1, "法术": 2, "真实": 3}
        damage_cfg = hit_cfg.get("damage") or {}
        d_elem = damage_cfg.get("element", "随技能")
        d_elem_int = elem_map.get(d_elem, 0) if isinstance(d_elem, str) else int(d_elem)
        d_type = damage_cfg.get("type", "随技能")
        d_type_int = type_map.get(d_type, 0) if isinstance(d_type, str) else int(d_type)

        # 表现模板（无特效时 model=0 也占位）
        present_node = make_template_call_node(
            "子弹通用逻辑-表现",
            extra_params=[0, 0, hit_cfg.get("hit_effect_model", 0), 2, 100, 2, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, False, 30, 0],
            alloc=alloc, ctx=ctx, desc="子弹通用-表现",
        )
        extra_nodes.append(present_node)

        # 伤害模板
        damage_node = make_template_call_node(
            "子弹通用逻辑-伤害",
            extra_params=[
                1, "entity:施法者伤害归属",
                d_type_int, d_elem_int,
                damage_cfg.get("subtype_flags", 16),
                damage_cfg.get("coef", 15000),
                damage_cfg.get("extra", -1),
                0, 0, 0,
                1 if hit_cfg.get("destroy_on_hit", False) else 0,
                -1, -1,
            ],
            alloc=alloc, ctx=ctx, desc="子弹通用-伤害",
        )
        extra_nodes.append(damage_node)

        # 命中 ORDER
        hit_order_node = Node(
            guid=alloc.new_guid(),
            cls="TSET_ORDER_EXECUTE",
            config_id=alloc.allocate_effect_id(),
            desc="命中后执行",
            config_payload={
                "ID": 0,
                "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
                "Params": [
                    TParam(value=damage_node.config_id).to_dict(),
                    TParam(value=present_node.config_id).to_dict(),
                ],
            },
            extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
            table_name="SkillEffectConfig",
        )
        hit_order_node.config_payload["ID"] = hit_order_node.config_id
        extra_nodes.append(hit_order_node)

        # 碰撞模板（圆形，命中范围内所有目标）
        # 注意：碰撞偏移是相对于子弹自身位置，而非施法者 — 子弹已在终点，偏移固定 0/0
        collision_node = make_template_call_node(
            "子弹通用逻辑-碰撞",
            extra_params=[
                hit_order_node.config_id,
                1,                                        # shape_type=1 圆形
                hit_cfg.get("collision_radius", 500),
                hit_cfg.get("collision_height", 360),
                0, 0, 0,                                  # 子弹自身为中心，偏移 0
                0,                                        # target_camp=0 敌方
                0, 1,                                     # 自定义条件, 跟随创建者
                hit_cfg.get("detect_interval", 1),
                hit_cfg.get("detect_cooldown", 0),
                0,
            ],
            alloc=alloc, ctx=ctx, desc=f"子弹通用-碰撞(圆形AOE R={hit_cfg.get('collision_radius', 500)})",
        )
        extra_nodes.append(collision_node)
        after_born_id = collision_node.config_id

    # BulletConfigNode：FlyType=0 静态子弹，LastTime=5帧
    bullet_node = _make_inline_bullet_config_node(
        bullet_id=bullet_id,
        fly_type=0,
        trace_path_type=0,
        trace_path_params=[],
        speed=0,
        last_time=5,
        max_distance=0,
        after_born_id=after_born_id,
        die_skill_id=0,
        alloc=alloc,
        ctx=ctx,
        desc="圆形AOE落点子弹",
    )

    # TSET_CREATE_BULLET
    create_node = Node(
        guid=alloc.new_guid(),
        cls="TSET_CREATE_BULLET",
        config_id=alloc.allocate_effect_id(),
        desc=cfg.get("desc", "创建圆形AOE子弹"),
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_BULLET"),
            "Params": [
                TParam(value=bullet_id).to_dict(),
                resolve_ref("attr:面向").to_dict(),
                resolve_ref("attr:位置X").to_dict(),
                resolve_ref("attr:位置Y").to_dict(),
                resolve_ref("entity:主体").to_dict(),
                TParam(value=offset_right).to_dict(),
                TParam(value=offset_forward).to_dict(),
                TParam(value=0).to_dict(),
                TParam(value=0).to_dict(),
                TParam(value=1 if cfg.get("is_projectile", True) else 0).to_dict(),
                TParam(value=0).to_dict(),
                TParam(value=0).to_dict(),
                TParam(value=0).to_dict(),
                TParam(value=0).to_dict(),
                TParam(value=1 if cfg.get("rapid_affect", False) else 0).to_dict(),
            ],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_CREATE_BULLET")},
        table_name="SkillEffectConfig",
    )
    create_node.config_payload["ID"] = create_node.config_id

    return [create_node, bullet_node] + extra_nodes


def expand_step(step: dict, alloc: IdAllocator, ctx: "BuildCtx") -> list[Node]:
    """单步 → Node 列表（首项是"代表节点"，可被父级引用）"""
    if "cast_anim" in step:
        return expand_cast_anim(step, alloc, ctx)
    if "cast_effect" in step:
        return expand_cast_effect(step, alloc, ctx)
    if "delay" in step:
        return expand_delay(step, alloc, ctx)
    if "bullet" in step:
        pat = step["bullet"].get("pattern", "直线子弹")
        if pat == "直线子弹":
            return expand_bullet_straight(step, alloc, ctx)
        if pat == "回旋飞回":
            return expand_bullet_boomerang(step, alloc, ctx)
        if pat == "圆形AOE":
            return expand_aoe_bullet_circle(step, alloc, ctx)
        if pat == "链状指示器":
            return expand_bullet_chain(step, alloc, ctx)
        raise CompileError(f"未实现的子弹模式: {pat}")
    if "play_sound" in step:
        return expand_play_sound(step, alloc, ctx)
    if "camera_shake" in step:
        return expand_camera_shake(step, alloc, ctx)
    if "apply_buff" in step:
        return expand_apply_buff(step, alloc, ctx)
    if "remove_buff" in step:
        return expand_remove_buff(step, alloc, ctx)
    if "modify_tag" in step:
        return expand_modify_tag(step, alloc, ctx)
    # v2.0 新增 step
    if "displacement" in step:
        return expand_displacement(step, alloc, ctx)
    if "summon_clone" in step:
        return expand_summon_clone(step, alloc, ctx)
    if "two_stage_skill" in step:
        return expand_two_stage_skill(step, alloc, ctx)
    if "return_to_clone" in step:
        return expand_return_to_clone(step, alloc, ctx)
    if "aoe_circle" in step:
        return expand_aoe_circle(step, alloc, ctx)
    raise CompileError(f"未识别的 flow step: keys={list(step.keys())}")


def build_order(steps: list[dict], alloc: IdAllocator, ctx: "BuildCtx",
                desc: str = "顺序执行") -> tuple[Node, list[Node]]:
    """把多个 step 编译为 TSET_ORDER_EXECUTE 节点 + 子节点。"""
    children_all: list[Node] = []
    children_root_ids: list[int] = []
    for s in steps:
        sub = expand_step(s, alloc, ctx)
        children_all.extend(sub)
        children_root_ids.append(sub[0].config_id)

    params = [TParam(value=cid) for cid in children_root_ids]
    order_node = Node(
        guid=alloc.new_guid(),
        cls="TSET_ORDER_EXECUTE",
        config_id=alloc.allocate_effect_id(),
        desc=desc,
        config_payload={
            "ID": 0,
            "SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE"),
            "Params": [p.to_dict() for p in params],
        },
        extra_data={"SkillEffectType": enum_to_int("TSkillEffectType", "TSET_ORDER_EXECUTE")},
        table_name="SkillEffectConfig",
    )
    order_node.config_payload["ID"] = order_node.config_id
    return order_node, children_all


# ------------------------------------------------------------
# 主节点构造
# ------------------------------------------------------------
def make_skill_config_node(ir: dict, root_effect_id: int, alloc: IdAllocator) -> Node:
    skill = ir["skill"]
    meta = ir["meta"]
    skill_id = meta["skill_id"]

    # 五行 / 主类型 / 子类型 / CD 类型 / 伤害类型 中文 → int
    element_int = cn_to_int("TElementsType", skill.get("element", "无")) or 0
    main_int    = cn_to_int("TBattleSkillMainType", skill.get("main_type", "-")) or 0
    sub_int     = cn_to_int("TBattleSkillSubType", skill.get("sub_type", "-")) or 0
    cd_int      = cn_to_int("TSkillColdType", skill.get("cd_type", "普通")) or 1
    # 伤害类型枚举使用项目内定义；PoC 阶段先简单数字映射
    dmg_map = {"-": 0, "直接伤害": 1, "物理": 1, "法术": 2, "真实": 3}
    damage_type_int = dmg_map.get(skill.get("damage_type", "-"), 0)

    # === 指示器 ===
    # IR 可在 skill.indicator 显式声明，否则按 flow 推断默认
    # ⚠️ TIndicatorType 真实枚举（PostMortem #014 修正）：
    #   1 = 无目标 / 2 = 单目标 / 3 = 直线 / 4 = 双圆 / 5 = 多向 / 6 = 双圆抓捕
    #   ※ 之前误用 TShapeType-like 映射（直线=2 错），导致 PoC 系列技能 indicator 全错
    indicator_cfg = skill.get("indicator", {})
    # v2.4.1（PostMortem #020）：补齐 TIRT_SECTOR=7（之前漏，"扇形"被错配为"多向"5）
    indicator_type_map = {
        "-": 0, "NULL": 0,
        "无目标": 1,
        "单目标": 2, "单体": 2,
        "直线": 3,
        "双圆": 4, "圆形": 4,    # 兼容历史"圆形"叫法
        "多向": 5,
        "双圆抓捕": 6,
        "扇形": 7,                # TIRT_SECTOR — param=[角度]
    }
    indicator_type = indicator_cfg.get("type", "")
    if isinstance(indicator_type, str):
        indicator_type_int = indicator_type_map.get(indicator_type, 0)
    else:
        indicator_type_int = int(indicator_type)
    # 自动推断：若 IR 没指明且 flow 中（含嵌套）有 bullet/aoe，给合理默认
    if indicator_type_int == 0 and not indicator_cfg:
        def _scan(steps: list) -> int:
            for step in steps or []:
                if not isinstance(step, dict):
                    continue
                if "bullet" in step:
                    return 3  # 直线（TIRT_LINE）
                if "aoe_circle" in step:
                    return 4  # 双圆（TIRT_DOUBLE_CIRCLE）
                # 递归 delay.then / if.then / if.else
                for k in ("delay", "if"):
                    blk = step.get(k, {})
                    for sub_key in ("then", "else"):
                        sub_steps = blk.get(sub_key) if isinstance(blk, dict) else None
                        r = _scan(sub_steps)
                        if r:
                            return r
            return 0
        indicator_type_int = _scan(ir.get("flow", []))
    # 指示器参数（按 TIndicatorType 真实枚举分别处理 — PostMortem #014 修正）
    # 已知规则（参考真实样本 + 30122003/30122001 等直线技能的 param=[range]）：
    #   1 无目标   → []
    #   2 单目标   → []
    #   3 直线     → [range]
    #   4 双圆     → [range]
    #   5 多向     → [range, ?]
    if "param" in indicator_cfg:
        indicator_param = indicator_cfg["param"]
        if not isinstance(indicator_param, list):
            indicator_param = [indicator_param]
    else:
        if indicator_type_int in (3, 4):     # 直线 / 双圆
            indicator_param = [skill.get("range", 0)]
        elif indicator_type_int == 5:        # 多向
            indicator_param = [skill.get("range", 0), 0]
        elif indicator_type_int == 7:        # 扇形 — param=[角度]，缺省 90 度
            indicator_param = [skill.get("indicator_angle", 90)]
        else:                                # 0/1/2 无 param
            indicator_param = []

    # === 智能施法 / 施法目标（MOBA 通用默认）===
    ai_cast_cfg = skill.get("ai_cast", {}) or {}
    DEFAULT_MONSTER_RANK_COND = [16, 15, 14, 13, 12, 11, 4, 5, 6, 0]
    DEFAULT_CAMP_COND = [2]  # 敌方
    smart_priority_map = {"距离最近": 3, "血量最低": 1, "血量最高": 2, "无": 0}
    smart_priority = ai_cast_cfg.get("target_priority", "距离最近")
    smart_priority_int = smart_priority_map.get(smart_priority, 3) if isinstance(smart_priority, str) else int(smart_priority)
    smart_camp = ai_cast_cfg.get("target_camp", DEFAULT_CAMP_COND)
    smart_rank = ai_cast_cfg.get("target_monster_rank", DEFAULT_MONSTER_RANK_COND)
    smart_cond_template = ai_cast_cfg.get("target_cond_template", 1)

    cast_target_cfg = skill.get("cast_target", {}) or {}
    cast_camp = cast_target_cfg.get("target_camp", DEFAULT_CAMP_COND)
    cast_rank = cast_target_cfg.get("target_monster_rank", DEFAULT_MONSTER_RANK_COND)
    cast_cond_template = cast_target_cfg.get("target_cond_template", 1)

    # ConfigJson 体（核心字段）
    payload = {
        "SkillXinfaType": 0,
        "ID": skill_id,
        "SkillNameKey": 0,
        "SkillDescKey": 0,
        "Icon": skill.get("icon", ""),
        "ElementType": element_int,
        "SkillSchoolResType": 0,
        "FeatureLabel": 0,
        "BDLabels": [],
        "SkillAITags": [],
        "DamageType": damage_type_int,
        "SkillMainType": main_int,
        "SkillSubType": sub_int,
        "MPCost": skill.get("mp_cost", 0),
        "HunLiValue": skill.get("hunli_value", 0),
        "SectXinfaEnergyValue": skill.get("xinfa_energy", 0),
        "SkillCastFrame": skill.get("cast_frame", 0),
        "SkillBufferStartFrame": skill.get("buffer_start_frame", 0),
        "SkillBufferFrame": skill.get("buffer_frame", 0),
        "SkillBaseDuration": skill.get("base_duration", 0),
        "SkillCastIsNotInterruptable": False,
        "IsSkillBufferFrameCanMove": False,
        "CdType": cd_int,
        "CdTime": skill.get("cd_frames", 0),
        "ComboCdList": [
            {
                "CDTime":            c.get("cd_time", 0),
                "CastFrame":         c.get("cast_frame", 0),
                "BufferFrame":       c.get("buffer_frame", 0),
                "BaseDuration":      c.get("base_duration", 0),
                "BufferStartFrame":  c.get("buffer_start_frame", 0),
            } for c in skill.get("combo_cd_list", []) or []
        ],
        "CDMaxStoreCount": 0,
        "SkillFixCdTime": 0,
        "SkillRange":    skill.get("range", 0),
        "SkillMinRange": skill.get("min_range", 0),
        "AISkillRange":  skill.get("ai_range", skill.get("range", 0)),
        "ExtraAlertRange": 0,
        "IsCloseChaseInAlertRange": True,
        "NeedTargetInRange": False,
        "IsSkillCastNoTargetInIdle": False,
        "Condition": 0,
        "AICastCondition": 0,
        "SkillEffectExecuteInfo": {
            "SelectConfigID": 0,
            "SkillEffectConfigID": root_effect_id,
        },
        "SkillEffectPassiveExecuteInfo": {
            "SelectConfigID": 0,
            "SkillEffectConfigID": 0,
        },
        "SkillEffectOnUnEquip": 0,
        "ChantCounterValuesList": [],
        "LGDamageValuesList": [],
        "SkillDamageTagsList": [],
        "SkillTagsList": [],
        "SkillTipsConditionSkillTagsList": [],
        "SkillGrowthDesc": "",
        "TalentKeyWord": [],
        "UseSkillSpeedDownValue": 0,
        "UseSkillSpeedDownTime": 0,
        "SkillIndicatorType": indicator_type_int,
        "SkillIndicatorParam": indicator_param,
        "SkillIndicatorParamTagConfigIds": [],
        "SkillIndicatorResParam": [0],
        "SkillIndicatorResParamTagConfigIds": [],
        "LockEntityAfterUseSkillDuration": 0,
        "LockEntityPosTypeAfterUseSkill": 0,
        "SkillEffectOnSkillCastInterrupt": 0,
        "UseSkillForbidUpdateFaceDir": False,
        "UseType": 0,
        "ButtonUpConfig": 0,
        "InterruptConfig": 0,
        "ReActiveConfig": 0,
        "SkillQuality": skill.get("quality", 1),
        "IsPassiveSkillHideCD": False,
        "IsPassiveSkillNotRunByCD": False,
        "SkillProperty": 0,
        "SkillRangeTagConfigId": 0,
        "SkillMinRangeTagConfigId": 0,
        "SmartCastTargetBasePriority": smart_priority_int,
        "SmartCastTargetCondTemplate": smart_cond_template,
        "SmartCastTargetMonsterRankCond": smart_rank,
        "SmartCastTargetCampCond": smart_camp,
        "SmartCastNoTargetIndicatorPos": 0,
        "SmartCastNoTargetCancelUse": False,
        "CastTargetCondTemplate": cast_cond_template,
        "CastTargetMonsterRankCond": cast_rank,
        "CastTargetCampCond": cast_camp,
        "IsHideContinueUseSkillHeadBar": False,
        "SkillPriority": 0,
        "EnhanceSkillBuffConfigID": 0,
        "StatisticsDamageAloneForSummon": False,
        "SkillNameEditor": skill.get("skill_name_text", meta.get("skill_name", "")),
        "SkillDescEditor": skill.get("skill_desc_text", ""),
    }

    n = Node(
        guid=alloc.new_guid(),
        cls="SkillConfigNode",
        config_id=skill_id,
        desc="",
        config_payload=payload,
        table_name="SkillConfig",
    )
    return n


def make_tag_config_node(tag: dict, alloc: IdAllocator) -> Node:
    payload = {
        "ID": tag["id"],
        "TagType": 0,
        "Desc": tag["name"],
        "NameKey": 0,
        "DefaultValue": tag.get("default", 0),
        "FinalValueEffectID": 0,
        "RetainWhenDie": tag.get("retain_when_die", False),
    }
    return Node(
        guid=alloc.new_guid(),
        cls="SkillTagsConfigNode",
        config_id=tag["id"],
        desc=tag["name"],
        config_payload=payload,
        table_name="SkillTagsConfig",
    )


# ------------------------------------------------------------
# Edges 推导
# ------------------------------------------------------------
# 动态端口节点：所有 Params 出边共用 "0" 锚点端口
# （详见技能节点字典 §2 节点字段约定，与真实样本 30122001 校验一致）
DYNAMIC_PORT_NODES = {
    "TSET_ORDER_EXECUTE",
    "TSET_NUM_MAX",
    "TSET_NUM_MIN",
    # 后续节点遇到再加
}


def derive_edges(nodes: list[Node]) -> list[dict]:
    """从节点的 ConfigJson Params 反推 edges（同 skill_graph_to_mermaid 的逻辑）

    特殊处理：
    - DYNAMIC_PORT_NODES（如 TSET_ORDER_EXECUTE）→ outputPortIdentifier 全部为 "0"
    - 其他节点 → outputPortIdentifier 用真实 Params 索引
    """
    # 按 ConfigID 索引（SkillConfig / SkillEffectConfig 等共用一个 ID 命名空间，
    # 用 (table_name, config_id) 作为复合 key 避免冲突）
    id_map: dict[int, list[Node]] = {}
    for n in nodes:
        id_map.setdefault(n.config_id, []).append(n)

    def best_target(target_id: int) -> Node | None:
        cands = id_map.get(target_id, [])
        if not cands:
            return None
        for c in cands:
            if c.cls != "RefConfigBaseNode":
                return c
        return cands[0]

    edges: list[dict] = []

    for n in nodes:
        if n.cls == "SkillConfigNode":
            sei = n.config_payload.get("SkillEffectExecuteInfo", {})
            for field_name, key in [
                ("SkillEffectExecuteInfo.SkillEffectConfigID", sei.get("SkillEffectConfigID", 0)),
                ("SkillEffectExecuteInfo.SelectConfigID",      sei.get("SelectConfigID", 0)),
            ]:
                if key:
                    target = best_target(key)
                    if target:
                        edges.append({
                            "GUID": str(uuid.uuid4()),
                            "inputNodeGUID":  target.guid,
                            "outputNodeGUID": n.guid,
                            "inputFieldName": "ID",
                            "outputFieldName": "PackedMembersOutput",
                            "inputPortIdentifier": "0",
                            "outputPortIdentifier": field_name,
                            "isVisible": True,
                        })
            continue
        if n.cls == "BulletConfigNode":
            # BulletConfigNode 有 3 个生命周期钩子，全部建边
            # （v1.2 修复 PostMortem #007：之前只建 AfterBorn，导致 Die 钩子节点视觉上孤立）
            for hook_field in (
                "BeforeBornSkillEffectExecuteInfo",
                "AfterBornSkillEffectExecuteInfo",
                "DieSkillEffectExecuteInfo",
            ):
                hook = n.config_payload.get(hook_field, {}) or {}
                hook_id = hook.get("SkillEffectConfigID", 0)
                if hook_id:
                    target = best_target(hook_id)
                    if target:
                        edges.append({
                            "GUID": str(uuid.uuid4()),
                            "inputNodeGUID":  target.guid,
                            "outputNodeGUID": n.guid,
                            "inputFieldName": "ID",
                            "outputFieldName": "PackedMembersOutput",
                            "inputPortIdentifier": "0",
                            "outputPortIdentifier": f"{hook_field}.SkillEffectConfigID",
                            "isVisible": True,
                        })
            continue
        if n.cls in {"RefConfigBaseNode", "ModelConfigNode", "SkillTagsConfigNode"}:
            continue

        # TSET_*/TSCT_*/TSKILLSELECT_*：扫 Params
        is_dynamic = n.cls in DYNAMIC_PORT_NODES
        params = n.config_payload.get("Params", [])
        for i, p in enumerate(params):
            if not isinstance(p, dict):
                continue
            v = p.get("Value", 0)
            pt = p.get("ParamType", 0)
            if pt in (1, 5, 6) or v == 0:
                continue
            target = best_target(v)
            if target is None or target.guid == n.guid:
                continue
            out_port = "0" if is_dynamic else str(i)
            edges.append({
                "GUID": str(uuid.uuid4()),
                "inputNodeGUID":  target.guid,
                "outputNodeGUID": n.guid,
                "inputFieldName": "ID",
                "outputFieldName": "PackedParamsOutput",
                "inputPortIdentifier": "0",
                "outputPortIdentifier": out_port,
                "isVisible": True,
            })
    return edges


# ------------------------------------------------------------
# TableTash（占位实现）
# ------------------------------------------------------------
# PoC 阶段：用 32 位固定占位 hash。Unity 编辑器 LoadGraph 时如果 hash 不匹配，
# 会自动按字段名兜底解析——只要 ConfigJson 字段名正确，加载不会失败。
# 后续阶段从 TableDR_CS 表 schema 提取真实 hash。
TABLE_TASH_PLACEHOLDER = "0CFA05568A66FEA1DF3BA6FE40DB7080"
TABLE_TASH_MAP = {
    # 与 30122001 真实样本一致的 hash
    "SkillEffectConfig": "0CFA05568A66FEA1DF3BA6FE40DB7080",
    "SkillConfig":       "E50E65EC1A6EC53E59E9B3A412153E83",
    "SkillTagsConfig":   "6A8A6883BDFDA1411BB2461E65CB2D9B",
    "SkillSelectConfig": "0CFA05568A66FEA1DF3BA6FE40DB7080",
    "SkillConditionConfig": "0CFA05568A66FEA1DF3BA6FE40DB7080",
    "BulletConfig":      "0CFA05568A66FEA1DF3BA6FE40DB7080",
    "ModelConfig":       "B1156BD6A276DA7A8E082421FCF0B7A9",
    "SkillInterruptConfig": "0CFA05568A66FEA1DF3BA6FE40DB7080",
}


def get_table_tash(table_name: str) -> str:
    return TABLE_TASH_MAP.get(table_name, TABLE_TASH_PLACEHOLDER)


# ------------------------------------------------------------
# 最终 JSON Emitter
# ------------------------------------------------------------
def serialize_node_data(n: Node, idx: int) -> dict:
    """节点序列化为 references.RefIds 数组中的一项"""
    data = {
        "GUID": n.guid,
        "computeOrder": idx,
        "position": {
            "serializedVersion": "2",
            "x": n.position_x,
            "y": n.position_y,
            "width": 300.0,
            "height": 200.0,
        },
        "expanded": False,
        "debug": False,
        "nodeLock": False,
        "visible": True,
        "hideChildNodes": False,
        "hidePos": {"x": 0.0, "y": 0.0},
        "hideCounter": 0,
        "ID": n.config_id,
        "Desc": n.desc,
        "IsTemplate": False,
        "TemplateFlags": 0,
        "TemplateParams": [],
        "TemplateParamsDesc": [],
        "TemplateParamsCustomAdd": False,
        "TableTash": get_table_tash(n.table_name),
        "ConfigJson": json.dumps(n.config_payload, ensure_ascii=False, separators=(",", ":")),
        "Config2ID": n.get_config2id() if n.cls != "RefConfigBaseNode" else None,
    }

    # 类型特有顶层字段
    data.update(n.extra_data)

    # RefConfigBaseNode 特殊
    if n.cls == "RefConfigBaseNode":
        data.pop("ConfigJson", None)
        data.pop("Config2ID", None)
        data["TableManagerName"] = "TableDR.SkillEffectConfigManager"
        data["ManualID"] = n.config_id

    return data


def _build_sticky_notes(ir: dict, nodes: list[Node]) -> list[dict]:
    """v2.4：自动给 SkillGraph 加 StickyNote — 含 IR meta + 关键节点说明。

    格式参考 30212009【木宗门】奇术_人阶_千叶散华，每个 stickyNote 含：
      position (x/y/width/height) + GUID + title + content

    注意：title/content 仅用 Unity 内置字体支持的字符（ASCII + 中文 + 常见标点），
    不用 emoji（📋 🔗 ⚠️ 等会渲染成方框）。
    """
    meta = ir.get("meta", {})
    skill = ir.get("skill", {})
    skill_id = meta.get("skill_id", 0)
    skill_name = meta.get("skill_name", "")
    description = meta.get("description", "")
    author = meta.get("author", "AI 编译器")
    ir_version = meta.get("ir_version", "?")

    # 1. 主说明 StickyNote — 技能概览
    main_lines = [
        f"【AI 自动生成】{skill_name} ({skill_id})",
        f"",
        f"IR 版本: {ir_version}",
        f"作者: {author}",
        f"",
        f"-- 技能描述 --",
        description or "(无 description 字段)",
        f"",
        f"-- 关键参数 --",
        f"元素: {skill.get('element', '?')} / 主类型: {skill.get('main_type', '?')} / 子类型: {skill.get('sub_type', '?')}",
        f"伤害类型: {skill.get('damage_type', '?')} / CD: {skill.get('cd_frames', '?')} 帧",
        f"指示器: {skill.get('indicator', {}).get('type', '?')} / 距离: {skill.get('range', '?')}",
        f"时长: cast={skill.get('cast_frame', '?')} bd={skill.get('base_duration', '?')} buffer={skill.get('buffer_frame', '?')}",
        f"",
        f"-- 节点统计 --",
        f"总节点: {len(nodes)}",
    ]

    cls_count: dict[str, int] = {}
    for n in nodes:
        cls_count[n.cls] = cls_count.get(n.cls, 0) + 1
    template_calls = cls_count.get("TSET_RUN_SKILL_EFFECT_TEMPLATE", 0)
    main_lines.append(f"模板调用: {template_calls}")

    main_note = {
        "position": {
            "serializedVersion": "2",
            "x": -2000.0,
            "y": -1800.0,
            "width": 380.0,
            "height": 420.0,
        },
        "GUID": str(uuid.uuid4()),
        "title": f"[概览] {skill_name}",
        "content": "\n".join(main_lines),
    }

    # 2. 双段式技能 / 跨技能依赖说明（如适用）
    extra_notes = []
    flow = ir.get("flow", []) or []

    # 找 two_stage_skill 信息
    for s in flow:
        if "two_stage_skill" in s:
            ts = s["two_stage_skill"]
            second_id = ts.get("second_stage_skill_id", skill_id + 1)
            content_lines = [
                f"【注意】两段式技能（v2.1+）",
                f"",
                f"本技能 = 主 SkillConfig {skill_id}",
                f"第二段独立 SkillConfig: {second_id}",
                f"",
                f"切槽位机制（44014633 模板）：",
                f"  第一段末尾自动调用，把 caster 当前槽位的技能 ID 替换为 {second_id}",
                f"  持续 {ts.get('auto_reset_frames', 600)} 帧后自动还原",
                f"",
                f"【Unity 实测前必须】",
                f"  1. 双 JSON 都用 SkillEditor 同步数据到 Table",
                f"  2. 第二段 CD 已设为 {ts.get('second_stage_cd_frames', skill.get('base_duration', 30))} 帧（不继承主 CD）",
            ]
            extra_notes.append({
                "position": {
                    "serializedVersion": "2",
                    "x": -2000.0,
                    "y": -1300.0,
                    "width": 360.0,
                    "height": 320.0,
                },
                "GUID": str(uuid.uuid4()),
                "title": "[关联] 两段式说明",
                "content": "\n".join(content_lines),
            })
            break

    # 3. SkillTag 跨技能说明（如有 tags 声明）
    tags_decl = ir.get("tags", []) or []
    if tags_decl:
        tag_lines = ["SkillTag 声明", ""]
        for t in tags_decl:
            tag_lines.append(f"  {t['id']} - {t['name']} (default={t.get('default', 0)})")
        tag_lines.extend([
            "",
            "【注意】实体级 tag (Param[1]=0/PT=0)：",
            "  写入和读取都用此模式 -> 跨技能透明读取",
            "  详见 PostMortem #019",
        ])
        extra_notes.append({
            "position": {
                "serializedVersion": "2",
                "x": -1600.0,
                "y": -1800.0,
                "width": 320.0,
                "height": 240.0,
            },
            "GUID": str(uuid.uuid4()),
            "title": "[Tag] 声明",
            "content": "\n".join(tag_lines),
        })

    return [main_note] + extra_notes


def emit_skill_graph(nodes: list[Node], output_path: Path, ir: dict) -> dict:
    edges = derive_edges(nodes)

    # 给每个节点排版位置（横向布局）
    for i, n in enumerate(nodes):
        n.position_x = -1500.0 + (i % 5) * 350
        n.position_y = -1200.0 + (i // 5) * 250

    refs = []
    for i, n in enumerate(nodes):
        rid = 1000 + i
        n.rid = rid
        refs.append({
            "rid": rid,
            "type": {
                "class": n.cls,
                "ns": "NodeEditor",
                "asm": "NodeEditor",
            },
            "data": serialize_node_data(n, i),
        })

    skill_graph = {
        "serializationData": {
            "SerializedFormat": 0,
            "SerializedBytes": [],
            "ReferencedUnityObjects": [],
            "SerializedBytesString": "",
            "Prefab": {"instanceID": 0},
            "PrefabModificationsReferencedUnityObjects": [],
            "PrefabModifications": [],
            "SerializationNodes": [],
        },
        "nodes": [{"rid": 1000 + i} for i in range(len(nodes))],
        "edges": edges,
        "groups": [],
        "stackNodes": [],
        "pinnedElements": [{
            "position": {"serializedVersion": "2", "x": 0.0, "y": 35.0, "width": 416.0, "height": 578.0},
            "opened": False,
            "editorType": {
                "serializedType": "NodeEditor.ConfigPinnedView, NodeEditor, Version=0.0.0.0, Culture=neutral, PublicKeyToken=null"
            }
        }],
        "exposedParameters": [],
        "serializedParameterList": [],
        "stickyNotes": _build_sticky_notes(ir, nodes),
        "curTab": 0,
        "path": str(output_path).replace("\\", "/"),
        "references": {
            "version": 2,
            "RefIds": refs,
        },
    }
    return skill_graph


# ------------------------------------------------------------
# 编译主流程
# ------------------------------------------------------------
@dataclass
class BuildCtx:
    """编译过程中传递的上下文"""
    ir: dict
    skill_id: int
    parent_skill_id: int | None = None  # v2.3：second_stage 编译时由 _split_two_stage 设为 main skill_id
    log: list[str] = field(default_factory=list)

    def info(self, msg: str):
        self.log.append(f"[INFO] {msg}")


class CompileError(Exception):
    pass


INT32_MAX = 2_147_483_647

# v2.4：缓存项目内已用的所有 skill_id（首次扫描后缓存）
_USED_SKILL_IDS_CACHE: set[int] | None = None


def _check_int32(value: int, label: str):
    """v2.3.1 防御：所有 tag/skill/effect ID 必须 ≤ int32 最大值，
    否则 Newtonsoft.Json 反序列化失败（实测 3014200301 → -1280766995）。"""
    if value > INT32_MAX or value < -INT32_MAX - 1:
        raise CompileError(
            f"{label} ID={value} 溢出 int32（最大 {INT32_MAX}）；"
            f"Newtonsoft.Json 会解析为负数导致节点反序列化失败。"
            f"建议：tag 用 skill_id*10+offset（最多 9 位）"
        )


def _scan_used_skill_ids() -> set[int]:
    """v2.4：扫所有 SkillGraph_<id>_*.json 文件名提取已用的 skill_id。

    缓存结果。文件名格式：SkillGraph_<digits>...json
    """
    global _USED_SKILL_IDS_CACHE
    if _USED_SKILL_IDS_CACHE is not None:
        return _USED_SKILL_IDS_CACHE

    scan_dirs = [
        PROJECT_ROOT / "Assets" / "Thirds" / "NodeEditor" / "SkillEditor" / "Saves" / "Jsons",
    ]
    used: set[int] = set()
    pat = re.compile(r"^SkillGraph_(\d+)")
    for d in scan_dirs:
        if not d.exists():
            continue
        for p in d.rglob("SkillGraph_*.json"):
            m = pat.match(p.name)
            if m:
                used.add(int(m.group(1)))
    _USED_SKILL_IDS_CACHE = used
    return used


def _check_skill_id_unique(skill_id: int, allow_self: bool, label: str = "meta.skill_id"):
    """v2.4：检查 skill_id 是否与项目内已有 SkillGraph_*.json 冲突。

    allow_self=True 时（修改已有技能场景），允许 skill_id 已存在但是同一个文件正在被覆盖。
    （编译器在落盘前阶段无法精确判断"是否同文件"，所以采用 allow_self 由调用方控制）
    """
    used = _scan_used_skill_ids()
    if skill_id in used and not allow_self:
        raise CompileError(
            f"{label}={skill_id} 与项目内已有 SkillGraph_*.json 冲突！"
            f"已用 ID 列表（前 5 个）：{sorted(used)[:5]}... 共 {len(used)} 个。"
            f"请换一个未使用的 ID。"
        )


def validate_ir(ir: dict, allow_overwrite: bool = False):
    """v2.4：allow_overwrite 控制 skill_id 冲突检查。

    True = 修改已有技能（允许 skill_id 已存在）
    False = 新建技能（skill_id 必须未使用过）
    """
    schema = json.loads(IR_SCHEMA_PATH.read_text(encoding="utf-8"))
    try:
        jsonschema.validate(ir, schema)
    except jsonschema.exceptions.ValidationError as e:
        raise CompileError(f"IR Schema 校验失败: {e.message}\n位置: {list(e.absolute_path)}")

    # v2.3.1 int32 防御
    skill_id = ir.get("meta", {}).get("skill_id", 0)
    _check_int32(skill_id, "meta.skill_id")
    # v2.4 skill_id 唯一性检查
    _check_skill_id_unique(skill_id, allow_self=allow_overwrite, label="meta.skill_id")
    for tag in ir.get("tags", []) or []:
        _check_int32(tag["id"], f"tags[{tag.get('name','?')}]")
    # 递归 step：record_position_to_tag / target_position_tag / second_stage_skill_id
    def _walk(steps):
        for s in steps:
            if "summon_clone" in s:
                rt = s["summon_clone"].get("record_position_to_tag", 0)
                if rt:
                    _check_int32(rt,     "summon_clone.record_position_to_tag")
                    _check_int32(rt + 1, "summon_clone.record_position_to_tag+1 (Y)")
            if "return_to_clone" in s:
                bt = s["return_to_clone"].get("target_position_tag", 0)
                if bt:
                    _check_int32(bt,     "return_to_clone.target_position_tag")
                    _check_int32(bt + 1, "return_to_clone.target_position_tag+1 (Y)")
            if "two_stage_skill" in s:
                ts = s["two_stage_skill"]
                if "second_stage_skill_id" in ts:
                    sid2 = ts["second_stage_skill_id"]
                    _check_int32(sid2, "two_stage_skill.second_stage_skill_id")
                    _check_skill_id_unique(sid2, allow_self=allow_overwrite,
                                           label="two_stage_skill.second_stage_skill_id")
                _walk(ts.get("first_stage", []))
                _walk(ts.get("second_stage", []))
            if "delay" in s:
                _walk(s["delay"].get("then", []))
    _walk(ir.get("flow", []))
    _walk(ir.get("passive", []))


def _compile_single(
    ir: dict,
    output_path: Path,
    parent_skill_id: int | None = None,
) -> tuple[dict, BuildCtx]:
    """单个 IR → 单个 SkillGraph JSON（不做多 SkillConfig 拆分）。

    parent_skill_id：v2.3 — second_stage 编译时由 _split_two_stage 传入主技能 ID，
    供 expand_return_to_clone 跨技能读 SkillTag 时使用。
    """
    skill_id = ir["meta"]["skill_id"]
    alloc = IdAllocator(skill_id)
    ctx = BuildCtx(ir=ir, skill_id=skill_id, parent_skill_id=parent_skill_id)

    flow = ir["flow"]
    if not flow:
        raise CompileError("flow 不能为空")

    if len(flow) == 1:
        children = expand_step(flow[0], alloc, ctx)
        root_effect_id = children[0].config_id
        all_effect_nodes = children
    else:
        order_node, others = build_order(flow, alloc, ctx, desc="主动入口")
        root_effect_id = order_node.config_id
        all_effect_nodes = [order_node] + others

    ctx.info(f"flow 编译完成：{len(all_effect_nodes)} 个 effect 节点，入口 ID={root_effect_id}")

    skill_node = make_skill_config_node(ir, root_effect_id, alloc)

    tag_nodes = []
    for tag in ir.get("tags", []) or []:
        tag_nodes.append(make_tag_config_node(tag, alloc))
    ctx.info(f"声明 tags: {len(tag_nodes)} 个")

    all_nodes = [skill_node] + all_effect_nodes + tag_nodes
    ctx.info(f"全部节点: {len(all_nodes)}")

    skill_graph = emit_skill_graph(all_nodes, output_path, ir)
    ctx.info(f"输出 JSON: {output_path}")

    return skill_graph, ctx


def _split_two_stage(ir: dict, output_path: Path) -> tuple[dict | None, Path | None]:
    """v2.1：检查 flow 是否含 two_stage_skill；若有，提取 second_stage 为独立 IR。

    返回 (second_ir, second_output_path)；无 two_stage 则 (None, None)。
    注意：不修改原 ir，second_stage 的 first_stage swap 调用由 expand_two_stage_skill
    在编译期处理（读 ctx.skill_id + cfg.second_stage_skill_id）。
    """
    flow = ir.get("flow", []) or []
    ts_step = next((s for s in flow if "two_stage_skill" in s), None)
    if ts_step is None:
        return None, None

    cfg = ts_step["two_stage_skill"]
    main_id = ir["meta"]["skill_id"]
    second_id = cfg.get("second_stage_skill_id", main_id + 1)
    second_name = cfg.get("second_stage_skill_name") or \
                  f"{ir['meta'].get('skill_name','')}_第二段"

    main_bd = ir["skill"].get("base_duration", 30)
    # v2.2 fix #1: 第二段 CD 默认 = base_duration（仅够走完动作；
    # 不能继承主技能 CD（如 360）否则再次按键被 CD 阻挡，参考 30221002 CdTime=18 vs 30221000 CdTime=210）
    second_cd = cfg.get("second_stage_cd_frames", main_bd)

    # 深拷贝主 IR 作为基底，仅修改 meta + flow + cd
    second_ir = copy.deepcopy(ir)
    second_ir["meta"]["skill_id"] = second_id
    second_ir["meta"]["skill_name"] = second_name
    second_ir["meta"]["description"] = (
        f"[{main_id} 第二段] " + (ir["meta"].get("description", "") or "")
    )
    second_ir["skill"]["cd_frames"] = second_cd
    # 收紧 bd / buffer 满足 cast≤bd≤cd
    if second_cd < main_bd:
        second_ir["skill"]["base_duration"] = second_cd
        second_ir["skill"]["buffer_frame"] = second_cd
    cast_f = ir["skill"].get("cast_frame", 0)
    if cast_f > second_cd:
        second_ir["skill"]["cast_frame"] = max(0, second_cd - 1)
    # second_stage 不需要再嵌套 two_stage_skill — 直接用 second_stage 的 step 作为 flow
    second_ir["flow"] = cfg["second_stage"]
    # v2.3：tags 是第一段独有的（声明 + 写入），第二段只跨技能读取，不重复声明
    second_ir.pop("tags", None)

    # 输出路径
    second_path = output_path.with_name(
        f"SkillGraph_{second_id}_{second_name}.json"
    )
    return second_ir, second_path


def compile_ir(
    ir: dict,
    output_path: Path,
    verbose: bool = False,
    allow_overwrite: bool = False,
) -> list[tuple[dict, BuildCtx, Path]]:
    """v2.1 多 SkillConfig 输出。

    返回 [(skill_graph, ctx, path), ...]。两段式技能会产出两条记录：
      [0] 主 SkillConfig（含 first_stage + 切槽位）
      [1] 第二段独立 SkillConfig（仅含 second_stage 的 step）

    向后兼容：单段技能返回单元素列表。
    """
    validate_ir(ir, allow_overwrite=allow_overwrite)

    second_ir, second_path = _split_two_stage(ir, output_path)

    main_graph, main_ctx = _compile_single(ir, output_path)
    results: list[tuple[dict, BuildCtx, Path]] = [(main_graph, main_ctx, output_path)]

    if second_ir is not None:
        # 第二段需经 schema 校验（meta.skill_id 已变 + flow 替换）
        validate_ir(second_ir, allow_overwrite=allow_overwrite)
        main_id = ir["meta"]["skill_id"]
        second_graph, second_ctx = _compile_single(second_ir, second_path, parent_skill_id=main_id)
        results.append((second_graph, second_ctx, second_path))

    return results


# ------------------------------------------------------------
# CLI
# ------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="SkillGraph IR → JSON 编译器")
    parser.add_argument("input", help="输入 .skill.yaml 路径")
    parser.add_argument("--out", "-o", help="输出 .json 路径（默认与输入同目录）")
    parser.add_argument("--dry-run", action="store_true", help="只编译不写文件")
    parser.add_argument("--overwrite", action="store_true",
                        help="允许 skill_id 与已有 SkillGraph_*.json 冲突（修改已有技能场景必加）")
    parser.add_argument("-v", "--verbose", action="store_true", help="详细日志")
    args = parser.parse_args()

    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"错误：文件不存在 {in_path}", file=sys.stderr)
        sys.exit(1)

    ir = yaml.safe_load(in_path.read_text(encoding="utf-8"))

    skill_id = ir.get("meta", {}).get("skill_id", "unknown")
    skill_name = ir.get("meta", {}).get("skill_name", "")
    if args.out:
        out_path = Path(args.out)
    else:
        out_path = in_path.with_name(f"SkillGraph_{skill_id}_{skill_name}.json")

    try:
        results = compile_ir(ir, out_path, verbose=args.verbose, allow_overwrite=args.overwrite)
    except CompileError as e:
        print(f"❌ 编译错误: {e}", file=sys.stderr)
        sys.exit(2)

    if args.verbose:
        for _, ctx, _ in results:
            for line in ctx.log:
                print(line, file=sys.stderr)

    if args.dry_run:
        for skill_graph, _, p in results:
            print(f"✓ Dry run 通过 [{p.name}]：{len(skill_graph['references']['RefIds'])} 个节点，"
                  f"{len(skill_graph['edges'])} 条边")
        return

    for skill_graph, _, p in results:
        p.write_text(
            json.dumps(skill_graph, ensure_ascii=False, indent=4),
            encoding="utf-8",
        )
        print(f"✓ 编译成功 → {p}")
        print(f"  节点数: {len(skill_graph['references']['RefIds'])}")
        print(f"  边数:   {len(skill_graph['edges'])}")
    if len(results) > 1:
        print(f"  共生成 {len(results)} 个 SkillConfig（v2.1 两段式拆分）")


if __name__ == "__main__":
    main()
